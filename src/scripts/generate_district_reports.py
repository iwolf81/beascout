#!/usr/bin/env python3
"""
District BeAScout Report Generation System

Generates professional district reports for Heart of New England Council:
- Quinapoxet District and Soaring Eagle District reports
- Excel spreadsheet format for easy sorting and analysis
- Supportive guidance approach for unit improvement

Usage:
    python scripts/generate_district_reports.py data/raw/all_units_*_scored.json
    python scripts/generate_district_reports.py data/raw/all_units_01720_scored.json --output-dir reports/
"""

import json
import sys
import glob
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import re


class DistrictReportGenerator:
    def __init__(self, key_three_file: str = "data/input/Key 3 08-22-2025.xlsx"):
        """Initialize with Key Three member data"""
        self.key_three_data = self._load_key_three_data(key_three_file)
        self.district_towns = self._load_district_towns()
        
    def _load_key_three_data(self, file_path: str) -> pd.DataFrame:
        """Load Key Three member data from Excel file"""
        try:
            # Read Excel file, skip header rows to get to actual data
            df = pd.read_excel(file_path, header=None, skiprows=3)
            print(f"Loaded {len(df)} Key Three member records")
            
            # Set proper column names based on the structure we know
            if len(df.columns) >= 11:
                df.columns = ['district', 'unit_id', 'fullname', 'address', 'citystate', 'zipcode', 
                             'email', 'phone', 'position', 'unitcommorgname', 'status']
                
                # Filter to only active members
                df = df[df['status'] == 'ACTIVE']
                
                return df
            else:
                print(f"Unexpected file structure - {len(df.columns)} columns found")
                return pd.DataFrame()
        except Exception as e:
            print(f"Error loading Key Three data: {e}", file=sys.stderr)
            return pd.DataFrame()
    
    def _load_district_towns(self) -> Dict[str, List[str]]:
        """Load district town assignments from centralized mapping"""
        try:
            # Add project root to path
            import os
            sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
            from mapping.district_mapping import TOWN_TO_DISTRICT
            
            # Group towns by district
            quinapoxet_towns = []
            soaring_eagle_towns = []
            
            for town, district in TOWN_TO_DISTRICT.items():
                if district == "Quinapoxet":
                    quinapoxet_towns.append(town)
                elif district == "Soaring Eagle":
                    soaring_eagle_towns.append(town)
            
            print(f"Loaded {len(quinapoxet_towns)} Quinapoxet towns, {len(soaring_eagle_towns)} Soaring Eagle towns")
            return {
                "Quinapoxet": quinapoxet_towns,
                "Soaring Eagle": soaring_eagle_towns
            }
            
        except Exception as e:
            print(f"Warning: Could not load HNE district data: {e}", file=sys.stderr)
            return {"Quinapoxet": [], "Soaring Eagle": []}
    
    def _classify_unit_district(self, unit: Dict[str, Any]) -> str:
        """Get district from unit data (pre-calculated) or classify based on town"""
        # Use pre-calculated district if available
        if 'district' in unit and unit['district']:
            return unit['district']
        
        # Fallback to calculating from town (for backward compatibility)
        unit_town = unit.get('unit_town', '').strip()
        
        if unit_town in self.district_towns["Quinapoxet"]:
            return "Quinapoxet"
        elif unit_town in self.district_towns["Soaring Eagle"]:
            return "Soaring Eagle"
        else:
            return "Unknown"
    
    def _get_unit_key_three(self, unit: Dict[str, Any]) -> List[Dict[str, str]]:
        """Get Key Three members for a unit"""
        if self.key_three_data.empty:
            return [{"name": "None", "position": "None", "email": "None", "phone": "None"}] * 3
        
        # Get unit identifier for matching
        primary_identifier = unit.get('primary_identifier', '')
        
        # Normalize identifier for matching
        identifier_for_matching = self._normalize_unit_identifier(primary_identifier)
        
        # Find matching Key Three members
        key_three_members = []
        
        for _, row in self.key_three_data.iterrows():
            org_name = str(row.get('unitcommorgname', '')).strip()
            normalized_org = self._normalize_unit_identifier(org_name)
            
            if normalized_org == identifier_for_matching:
                member_data = {
                    'name': str(row.get('fullname', 'Missing')).strip(),
                    'position': str(row.get('position', 'Missing')).strip(), 
                    'email': str(row.get('email', 'Missing')).strip(),
                    'phone': str(row.get('phone', 'Missing')).strip()
                }
                
                # Clean up missing data
                for key, value in member_data.items():
                    if value in ['nan', 'None', '', 'NaT']:
                        member_data[key] = 'None'
                
                key_three_members.append(member_data)
        
        # Enhanced matching for missing units (Key Three Only)
        if len(key_three_members) == 0 and unit.get('data_source') == 'Key Three Only':
            # Try more flexible matching for missing units
            unit_type = unit.get('unit_type', '')
            unit_number = unit.get('unit_number', '').lstrip('0')
            unit_town = unit.get('unit_town', '')
            
            for _, row in self.key_three_data.iterrows():
                org_name = str(row.get('unitcommorgname', '')).strip()
                
                # Check if unit type, number, and town all appear in the org name
                if (unit_type.lower() in org_name.lower() and 
                    unit_number in org_name and
                    unit_town.lower() in org_name.lower()):
                    
                    member_data = {
                        'name': str(row.get('fullname', 'Missing')).strip(),
                        'position': str(row.get('position', 'Missing')).strip(), 
                        'email': str(row.get('email', 'Missing')).strip(),
                        'phone': str(row.get('phone', 'Missing')).strip()
                    }
                    
                    # Clean up missing data
                    for key, value in member_data.items():
                        if value in ['nan', 'None', '', 'NaT']:
                            member_data[key] = 'None'
                    
                    key_three_members.append(member_data)
        
        # Ensure we have exactly 3 members (pad with "None" if needed)
        while len(key_three_members) < 3:
            key_three_members.append({
                "name": "None", 
                "position": "None",
                "email": "None", 
                "phone": "None"
            })
        
        return key_three_members[:3]  # Take only first 3
    
    def _normalize_unit_identifier(self, identifier: str) -> str:
        """Normalize unit identifier for matching"""
        if not identifier:
            return ""
        
        # Remove extra spaces and standardize
        normalized = re.sub(r'\s+', ' ', identifier.strip())
        
        # Remove gender indicators like "(F)", "(B)", "(G)" and associated dashes/spaces
        normalized = re.sub(r'\s*\([FBG]\)\s*-?\s*', ' ', normalized)
        
        # Remove Crew specialty information (everything after "Specialty:")
        if 'Specialty:' in normalized:
            normalized = normalized.split('Specialty:')[0].strip()
        
        # Clean up multiple spaces that may result from the above
        normalized = re.sub(r'\s+', ' ', normalized).strip()
        
        # Ensure consistent format: "Type Number Organization"
        # This handles cases where there might be inconsistent spacing or punctuation
        parts = normalized.split()
        if len(parts) >= 2:
            unit_type = parts[0]
            unit_number = parts[1].lstrip('0')  # Remove leading zeros for comparison
            
            # Handle organization name - everything after unit number
            if len(parts) > 2:
                # Join all remaining parts and then clean
                org_name = ' '.join(parts[2:])
                
                # Remove leading dash and spaces that might result from "Type Number - OrgName" format
                org_name = org_name.lstrip('- ').strip()
                
                # Standardize common variations - remove punctuation and normalize spaces
                org_name = org_name.replace(',', '').replace('.', '').replace('  ', ' ')
                org_name = org_name.replace(' Inc', '').replace(' Inc.', '')  # Handle "Inc" variations
                
                # Normalize dashes - standardize spacing around dashes
                org_name = re.sub(r'\s*-\s*', '-', org_name)  # Remove spaces around dashes
                
                org_name = org_name.strip()
                
                return f"{unit_type} {unit_number} {org_name}"
            else:
                return f"{unit_type} {unit_number}"
        
        return normalized
    
    def _categorize_issues(self, recommendations: List[str]) -> Dict[str, List[str]]:
        """Categorize issues into Required, Quality, Recommended"""
        required_issues = []
        quality_issues = []
        recommended_issues = []
        
        for rec in recommendations:
            if rec.startswith('REQUIRED_MISSING_'):
                issue_name = rec.replace('REQUIRED_MISSING_', '').replace('_', ' ').title()
                required_issues.append(issue_name)
            elif rec.startswith('QUALITY_'):
                issue_name = rec.replace('QUALITY_', '').replace('_', ' ').title()
                quality_issues.append(issue_name)
            elif rec.startswith('RECOMMENDED_MISSING_'):
                issue_name = rec.replace('RECOMMENDED_MISSING_', '').replace('_', ' ').title()
                recommended_issues.append(issue_name)
        
        return {
            'required': required_issues,
            'quality': quality_issues, 
            'recommended': recommended_issues
        }
    
    def _create_unit_data_headers(self, units: List[Dict[str, Any]], district_name: str) -> List[str]:
        """Create header rows for Unit Data sheet matching user specification"""
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M%p").lower()
        
        # Calculate metrics
        total_units = len(units)
        towns = set(unit.get('unit_town', '') for unit in units)
        town_count = len(towns)
        
        # Grade distribution - count F & D grades for outreach
        outreach_needed = 0
        personal_email_count = 0
        
        for unit in units:
            grade = unit.get('completeness_grade', '')
            if grade in ['F', 'D']:
                outreach_needed += 1
            
            recommendations = unit.get('recommendations', [])
            if 'QUALITY_PERSONAL_EMAIL' in recommendations:
                personal_email_count += 1
        
        # Create header rows exactly as specified
        return [
            "BeAScout Quality Report",
            f"{district_name} District - Heart of New England Council (HNE), Scouting America", 
            f"Generated: {date_str} {time_str}",
            "",
            "Generated by HNE Council Executive Board for Unit Improvement Initiatives.",
            "This information is to be used only for authorized purposes on behalf of Scouting America.",
            "Disclosing, copying, or making any inappropriate use of this information is strictly prohibited.",
            "",
            'Data Sources: BeAScout.org (10-mile radius) + JoinExploring.org (20-mile radius)',
            f"Last Complete Data Retrieval: {date_str}",
            "",
            f"District Coverage: {total_units} units analyzed across {town_count} towns",
            f"Key Three outreach required (F & D grades): {outreach_needed}",
            f"Units using personal emails: {personal_email_count}",
            "",
            "Legend:",
            "Location - Missing meeting location (street address required for families to find meetings)",
            "Day - Missing meeting day (families need to know when meetings occur)",
            "Time - Missing meeting time (specific meeting times help families plan attendance)",
            "Personal Email - Unit-specific contact email is preferable (personal emails create continuity issues when leaders change)",
            "Website - Recommend a unit-specific website (provides families detailed program information)",
            "Contact - Recommend a contact name (gives families a specific person to reach out to)",
            "Phone - Recommend a contact phone number (provides direct communication method for interested families)",
            ""
        ]
    
    def _write_unit_data_with_headers(self, writer, header_rows: List[str], df_main: pd.DataFrame, sheet_name: str):
        """Write header rows followed by unit data to Excel sheet with formatting"""
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font
        
        # Create worksheet
        ws = writer.book.create_sheet(sheet_name)
        
        # Write header rows
        for row_num, header_text in enumerate(header_rows, 1):
            ws[f'A{row_num}'] = header_text
        
        # Apply formatting to specific header rows
        bold_font = Font(bold=True)
        
        # Row 1: Report Name (bold)
        ws['A1'].font = bold_font
        
        # Row 2: District name (bold) 
        ws['A2'].font = bold_font
        
        # Row 12: District Coverage (bold) - find the row with "District Coverage"
        for row_num, header_text in enumerate(header_rows, 1):
            if "District Coverage:" in str(header_text):
                ws[f'A{row_num}'].font = bold_font
                break
        
        # Write unit data starting after headers
        start_row = len(header_rows) + 1
        
        # Write column headers (bold)
        for col_num, col_name in enumerate(df_main.columns, 1):
            cell = ws.cell(row=start_row, column=col_num, value=col_name)
            cell.font = bold_font
        
        # Write data rows
        for r_idx, row in enumerate(dataframe_to_rows(df_main, index=False, header=False), start_row + 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Set column widths
        # Columns A-E: width 15
        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col_letter].width = 15
        
        # Columns F and beyond: width 20
        for col_num in range(6, len(df_main.columns) + 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 20
    
    def _create_header_data_legacy(self, units: List[Dict[str, Any]], district_name: str) -> List[Dict[str, str]]:
        """Create professional header data based on user requirements"""
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M%p").lower()
        
        # Calculate metrics
        total_units = len(units)
        towns = set(unit.get('unit_town', '') for unit in units)
        town_count = len(towns)
        
        # Grade distribution
        grade_counts = {'F': 0, 'D': 0}
        personal_email_count = 0
        
        for unit in units:
            grade = unit.get('completeness_grade', '')
            if grade in ['F', 'D']:
                grade_counts[grade] += 1
            
            recommendations = unit.get('recommendations', [])
            if 'QUALITY_PERSONAL_EMAIL' in recommendations:
                personal_email_count += 1
        
        outreach_needed = grade_counts['F'] + grade_counts['D']
        
        return [
            {"Header": "BeAScout Quality Report"},
            {"Header": "Heart of New England Council (HNE), Scouting America"},
            {"Header": f"Generated: {date_str} {time_str}"},
            {"Header": ""},
            {"Header": "Generated by HNE Council Executive Board for Unit Improvement Initiatives."},
            {"Header": "This information is to be used only for authorized purposes on behalf of Scouting America."},
            {"Header": "Disclosing, copying, or making any inappropriate use of this information is strictly prohibited."},
            {"Header": ""},
            {"Header": 'Data Sources: BeAScout.org (10-mile radius) + JoinExploring.org (20-mile radius)'},
            {"Header": f"Last Complete Data Retrieval: {date_str}"},
            {"Header": ""},
            {"Header": f"District Coverage: {total_units} units analyzed across {town_count} towns"},
            {"Header": f"Key Three outreach required (F & D grades): {outreach_needed}"},
            {"Header": f"Units using personal emails: {personal_email_count}"},
            {"Header": ""}
        ]
    
    def _create_expanded_legend_data(self) -> List[Dict[str, str]]:
        """Create expanded legend data with user-requested clarity"""
        return [
            {"Category": "Required Information", "Issue": "Location", "Description": "Missing meeting location (street address required for families to find meetings)"},
            {"Category": "Required Information", "Issue": "Day", "Description": "Missing meeting day (families need to know when meetings occur)"},
            {"Category": "Required Information", "Issue": "Time", "Description": "Missing meeting time (specific meeting times help families plan attendance)"},
            {"Category": "Required Information", "Issue": "Contact Email", "Description": "Missing contact email (families need a way to ask questions and get information)"},
            {"Category": "Required Information", "Issue": "Specialty", "Description": "Missing specialty focus (required for Venturing Crews - e.g., High Adventure, STEM)"},
            {"Category": "Quality Issues", "Issue": "Personal Email", "Description": "Unit-specific contact email is preferable (personal emails create continuity issues when leaders change)"},
            {"Category": "Quality Issues", "Issue": "PO Box Location", "Description": "Street address preferred over PO Box (helps families locate actual meeting place)"},
            {"Category": "Recommended Information", "Issue": "Contact", "Description": "Recommend a contact name (gives families a specific person to reach out to)"},
            {"Category": "Recommended Information", "Issue": "Phone", "Description": "Recommend a contact phone number (provides direct communication method for interested families)"},
            {"Category": "Recommended Information", "Issue": "Website", "Description": "Recommend a unit-specific website (provides families detailed program information)"},
            {"Category": "Recommended Information", "Issue": "Description", "Description": "Recommend program description (helps families understand what makes the unit special and welcoming)"}
        ]
    
    def _create_legacy_legend_data(self) -> List[Dict[str, str]]:
        """Legacy legend method for backwards compatibility"""
        return [
            {"Code": "Meeting Location", "Category": "Required", "Description": "Street address where unit meets - helps families find meetings and events"},
            {"Code": "Meeting Day", "Category": "Required", "Description": "Day of week when unit regularly meets - essential for family scheduling"},
            {"Code": "Meeting Time", "Category": "Required", "Description": "Time when meetings start and end - allows families to plan attendance"},
            {"Code": "Contact Email", "Category": "Required", "Description": "Email address for families to ask questions and get information"},
            {"Code": "Specialty", "Category": "Required", "Description": "Program focus area for Venturing Crews (e.g., High Adventure, STEM)"},
            {"Code": "Personal Email", "Category": "Quality", "Description": "Consider using unit-specific email address for better continuity as leaders change"},
            {"Code": "Po Box Location", "Category": "Quality", "Description": "Street address preferred over PO Box - helps families locate actual meeting place"},
            {"Code": "Contact Person", "Category": "Recommended", "Description": "Designated contact person's name - gives families specific person to reach out to"},
            {"Code": "Phone Number", "Category": "Recommended", "Description": "Phone contact for immediate questions and information"},
            {"Code": "Website", "Category": "Recommended", "Description": "Unit website with additional program information and updates"},
            {"Code": "Description", "Category": "Recommended", "Description": "Program description highlighting what makes the unit special and welcoming"}
        ]
    
    def generate_district_report(self, units: List[Dict[str, Any]], district_name: str, output_dir: str) -> str:
        """Generate Excel report for a specific district"""
        
        # Filter units for this district
        district_units = [unit for unit in units if self._classify_unit_district(unit) == district_name]
        
        if not district_units:
            print(f"No units found for {district_name} District")
            return ""
        
        # Sort units by type, number, town
        district_units.sort(key=lambda x: (
            x.get('unit_type', ''),
            int(x.get('unit_number', '0') or '0'),
            x.get('unit_town', '')
        ))
        
        # Prepare data for Excel
        report_data = []
        
        for unit in district_units:
            # Get Key Three members
            key_three_members = self._get_unit_key_three(unit)
            
            # Categorize issues
            recommendations = unit.get('recommendations', [])
            issues = self._categorize_issues(recommendations)
            
            # Format issues as comma-separated strings with 'None' for empty
            required_str = ', '.join(issues['required']) if issues['required'] else 'None'
            quality_str = ', '.join(issues['quality']) if issues['quality'] else 'None'
            recommended_str = ', '.join(issues['recommended']) if issues['recommended'] else 'None'
            
            # Build row data
            row_data = {
                'Unit Type': unit.get('unit_type', ''),
                'Unit Number': unit.get('unit_number', ''),
                'Town': unit.get('unit_town', ''),
                'BeAScout Score (%)': round(unit.get('completeness_score', 0), 1),
                'Grade': unit.get('completeness_grade', ''),
                'Missing Information': required_str,
                'Quality Issues': quality_str,
                'Recommended Information': recommended_str,
            }
            
            # Add Key Three members
            for i, member in enumerate(key_three_members, 1):
                row_data[f'Key Three #{i} Name'] = member['name']
                row_data[f'Key Three #{i} Position'] = member['position']
                row_data[f'Key Three #{i} Email'] = member['email']
                row_data[f'Key Three #{i} Phone'] = member['phone']
            
            report_data.append(row_data)
        
        # Create Excel file
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        date_str = datetime.now().strftime("%Y%m%d")
        time_str = datetime.now().strftime("%H%M%S")
        filename = f"{district_name}_District_BeAScout_Report_{date_str}_{time_str}.xlsx"
        file_path = output_path / filename
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Unit Data sheet with header integrated at top
            header_rows = self._create_unit_data_headers(district_units, district_name)
            df_main = pd.DataFrame(report_data)
            
            # Write header rows first, then unit data
            self._write_unit_data_with_headers(writer, header_rows, df_main, 'Unit Data')
            
            # Summary sheet (keep this)
            summary_data = self._create_summary_data(district_units, district_name)
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='District Summary', index=False)
        
        print(f"Generated: {filename}")
        return str(file_path)
    
    def _create_summary_data(self, units: List[Dict[str, Any]], district_name: str) -> List[Dict[str, Any]]:
        """Create summary statistics for the district"""
        if not units:
            return []
        
        total_units = len(units)
        scores = [unit.get('completeness_score', 0) for unit in units]
        avg_score = sum(scores) / len(scores) if scores else 0
        
        # Grade distribution
        grade_counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
        for score in scores:
            if score >= 90:
                grade_counts['A'] += 1
            elif score >= 80:
                grade_counts['B'] += 1
            elif score >= 70:
                grade_counts['C'] += 1
            elif score >= 60:
                grade_counts['D'] += 1
            else:
                grade_counts['F'] += 1
        
        # Unit type distribution
        unit_types = {}
        for unit in units:
            unit_type = unit.get('unit_type', 'Unknown')
            unit_types[unit_type] = unit_types.get(unit_type, 0) + 1
        
        summary = [
            {"Metric": "District", "Value": f"{district_name} District"},
            {"Metric": "Report Generated", "Value": datetime.now().strftime("%Y-%m-%d %H:%M")},
            {"Metric": "", "Value": ""},
            {"Metric": "OVERVIEW", "Value": ""},
            {"Metric": "Total Units", "Value": total_units},
            {"Metric": "Average BeAScout Score", "Value": f"{avg_score:.1f}%"},
            {"Metric": "", "Value": ""},
            {"Metric": "GRADE DISTRIBUTION", "Value": ""},
            {"Metric": "A (90-100%)", "Value": f"{grade_counts['A']} units ({grade_counts['A']/total_units*100:.1f}%)"},
            {"Metric": "B (80-89%)", "Value": f"{grade_counts['B']} units ({grade_counts['B']/total_units*100:.1f}%)"},
            {"Metric": "C (70-79%)", "Value": f"{grade_counts['C']} units ({grade_counts['C']/total_units*100:.1f}%)"},
            {"Metric": "D (60-69%)", "Value": f"{grade_counts['D']} units ({grade_counts['D']/total_units*100:.1f}%)"},
            {"Metric": "F (<60%)", "Value": f"{grade_counts['F']} units ({grade_counts['F']/total_units*100:.1f}%)"},
            {"Metric": "", "Value": ""},
            {"Metric": "UNIT TYPES", "Value": ""},
        ]
        
        for unit_type, count in sorted(unit_types.items()):
            summary.append({"Metric": f"{unit_type}s", "Value": f"{count} units"})
        
        return summary
    
    def generate_reports_from_files(self, json_files: List[str], output_dir: str = "data/output/reports/") -> List[str]:
        """Generate district reports from multiple JSON files"""
        all_units = []
        
        # Load all units from all files
        for file_path in json_files:
            try:
                with open(file_path, 'r') as f:
                    data = json.load(f)
                
                units = data.get('units_with_scores', [])
                
                # If this is the authoritative final dataset, trust it's already HNE-only
                if 'authoritative_final' in file_path:
                    hne_units = units  # Trust the authoritative dataset 
                    print(f"Using authoritative dataset - no additional filtering needed")
                else:
                    # Filter to only HNE Council units for other datasets
                    hne_units = []
                    for unit in units:
                        unit_town = unit.get('unit_town', '')
                        if (unit_town in self.district_towns["Quinapoxet"] or 
                            unit_town in self.district_towns["Soaring Eagle"]):
                            hne_units.append(unit)
                
                all_units.extend(hne_units)
                print(f"Loaded {len(hne_units)} HNE units from {file_path}")
                
            except Exception as e:
                print(f"Error loading {file_path}: {e}", file=sys.stderr)
                continue
        
        if not all_units:
            print("No HNE Council units found to process")
            return []
        
        print(f"Total HNE units loaded: {len(all_units)}")
        
        # Generate single council report with both district sheets
        council_report = self.generate_council_report(all_units, output_dir)
        return [council_report] if council_report else []
    
    def generate_council_report(self, all_units: List[Dict[str, Any]], output_dir: str) -> str:
        """Generate single council report with separate district sheets"""
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        date_str = datetime.now().strftime("%Y%m%d")
        time_str = datetime.now().strftime("%H%M%S")
        filename = f"HNE_Council_BeAScout_Report_{date_str}_{time_str}.xlsx"
        file_path = output_path / filename
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Generate both district sheets
            for district_name in ["Quinapoxet", "Soaring Eagle"]:
                district_units = [unit for unit in all_units if self._classify_unit_district(unit) == district_name]
                
                # Always create sheet even if no units (per feedback requirement #1)
                if not district_units:
                    print(f"No units found for {district_name} District - creating empty sheet")
                
                # Sort units by type, number, town
                district_units.sort(key=lambda x: (
                    x.get('unit_type', ''),
                    int(x.get('unit_number', '0') or '0'),
                    x.get('unit_town', '')
                ))
                
                # Create header rows for this district
                header_rows = self._create_unit_data_headers(district_units, district_name)
                
                # Prepare unit data
                report_data = []
                for unit in district_units:
                    # Get Key Three members (up to 3)
                    key_three_members = self._get_unit_key_three(unit)[:3]
                    
                    # Categorize issues
                    recommendations = unit.get('recommendations', [])
                    issues = self._categorize_issues(recommendations)
                    
                    # Format issues as comma-separated strings
                    required_str = ', '.join(issues['required']) if issues['required'] else 'None'
                    quality_str = ', '.join(issues['quality']) if issues['quality'] else 'None'
                    recommended_str = ', '.join(issues['recommended']) if issues['recommended'] else 'None'
                    
                    # Build row data
                    row_data = {
                        'Unit Type': unit.get('unit_type', ''),
                        'Unit Number': unit.get('unit_number', ''),
                        'Town': unit.get('unit_town', ''),
                        'BeAScout Score (%)': round(unit.get('completeness_score', 0), 1),
                        'Grade': unit.get('completeness_grade', ''),
                        'Missing Information': required_str,
                        'Quality Issues': quality_str,
                        'Recommended Information': recommended_str,
                    }
                    
                    # Add Key Three members (exactly 3 slots)
                    for i in range(3):
                        if i < len(key_three_members):
                            member = key_three_members[i]
                            row_data[f'Key Three #{i+1} Name'] = member['name']
                            row_data[f'Key Three #{i+1} Position'] = member['position']
                            row_data[f'Key Three #{i+1} Email'] = member['email']
                            row_data[f'Key Three #{i+1} Phone'] = member['phone']
                        else:
                            # Empty slots for missing Key Three members
                            row_data[f'Key Three #{i+1} Name'] = ''
                            row_data[f'Key Three #{i+1} Position'] = ''
                            row_data[f'Key Three #{i+1} Email'] = ''
                            row_data[f'Key Three #{i+1} Phone'] = ''
                    
                    report_data.append(row_data)
                
                # Create dataframe and write to sheet
                df_main = pd.DataFrame(report_data)
                self._write_unit_data_with_headers(writer, header_rows, df_main, f'{district_name} District')
                
                print(f"Created {district_name} District sheet with {len(district_units)} units")
        
        print(f"Generated council report: {filename}")
        return str(file_path)


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/generate_district_reports.py <json_file_pattern> [--output-dir output/]", file=sys.stderr)
        print("Examples:")
        print("  python scripts/generate_district_reports.py data/raw/all_units_01720_scored.json")
        print("  python scripts/generate_district_reports.py data/raw/all_units_*_scored.json --output-dir reports/")
        sys.exit(1)
    
    # Parse arguments
    args = sys.argv[1:]
    output_dir = "data/output/reports/"
    
    if "--output-dir" in args:
        output_idx = args.index("--output-dir")
        if output_idx + 1 < len(args):
            output_dir = args[output_idx + 1]
            args = args[:output_idx] + args[output_idx + 2:]
    
    # Handle glob patterns
    all_files = []
    for pattern in args:
        if '*' in pattern or '?' in pattern:
            all_files.extend(glob.glob(pattern))
        else:
            all_files.append(pattern)
    
    # Remove duplicates and check file existence
    all_files = list(set(all_files))
    valid_files = []
    for file_path in all_files:
        if Path(file_path).exists():
            valid_files.append(file_path)
        else:
            print(f"File not found: {file_path}", file=sys.stderr)
    
    if not valid_files:
        print("No valid files found to process", file=sys.stderr)
        sys.exit(1)
    
    # Generate reports
    generator = DistrictReportGenerator()
    generated_files = generator.generate_reports_from_files(valid_files, output_dir)
    
    print(f"\nGenerated {len(generated_files)} district reports in {output_dir}")
    for file_path in generated_files:
        print(f"  {Path(file_path).name}")


if __name__ == "__main__":
    main()