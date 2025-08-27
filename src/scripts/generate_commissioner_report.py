#!/usr/bin/env python3
"""
Commissioner Data Quality Report Generator
Creates comprehensive Excel reports with validation results and action flags
Designed for district commissioners to address unit discrepancies
"""

import json
import sys
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Add project root to path for imports
sys.path.append(str(Path(__file__).parent.parent))

class CommissionerReportGenerator:
    """
    Generates comprehensive Excel reports for district commissioners
    Highlights units needing attention with clear action flags
    """
    
    def __init__(self):
        self.validation_data = None
        
    def load_validation_data(self, file_path: str = 'data/output/enhanced_three_way_validation_results.json') -> bool:
        """Load three-way validation results"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                self.validation_data = json.load(f)
                print(f"📊 Loaded validation data: {len(self.validation_data['validation_results'])} units")
                return True
        except Exception as e:
            print(f"❌ Failed to load validation data: {e}")
            return False
    
    def create_commissioner_report(self, output_path: str = None) -> str:
        """
        Create comprehensive commissioner report with action flags
        
        Returns:
            Path to generated Excel file
        """
        if not self.validation_data:
            raise ValueError("No validation data loaded")
        
        # Generate timestamped filename if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"data/output/reports/HNE_Commissioner_Data_Quality_Report_{timestamp}.xlsx"
        
        # Ensure output directory exists
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook with multiple sheets
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create report sheets
        self._create_executive_summary_sheet(workbook)
        self._create_key_three_only_sheet(workbook)
        self._create_web_only_sheet(workbook)
        self._create_data_quality_sheet(workbook)
        self._create_complete_unit_listing_sheet(workbook)
        
        # Save workbook
        workbook.save(output_path)
        print(f"\n📋 Commissioner report saved: {output_path}")
        
        return output_path
    
    def _create_executive_summary_sheet(self, workbook):
        """Create executive summary sheet with key metrics"""
        ws = workbook.create_sheet("Executive Summary")
        
        summary = self.validation_data['validation_summary']
        
        # Title
        ws['A1'] = "Heart of New England Council - Enhanced Data Quality Audit"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Report date and version
        ws['A2'] = f"Report Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A2'].font = Font(size=12)
        
        ws['A3'] = "Enhanced Version: Corrected town matching and territorial filtering applied"
        ws['A3'].font = Font(size=10, italic=True)
        
        # Key metrics
        row = 5
        ws[f'A{row}'] = "VALIDATION OVERVIEW"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 2
        metrics = [
            ("Total Units Analyzed", summary['total_units']),
            ("Units with Both Sources", f"{summary['status_breakdown']['both_sources']} ({summary['validation_percentages']['both_sources']:.1f}%)"),
            ("Units Missing from Web", f"{summary['status_breakdown']['key_three_only']} ({summary['validation_percentages']['key_three_only']:.1f}%)"),
            ("Units on Web Only", f"{summary['status_breakdown']['web_only']} ({summary['validation_percentages']['web_only']:.1f}%)"),
            ("Units with Data Quality Issues", summary['units_with_issues'])
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # District breakdown
        row += 2
        ws[f'A{row}'] = "UNITS MISSING FROM WEB BY DISTRICT"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 2
        for district, count in summary['key_three_only_by_district'].items():
            ws[f'A{row}'] = district
            ws[f'B{row}'] = f"{count} units"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Action summary
        row += 2
        ws[f'A{row}'] = "COMMISSIONER ACTION REQUIRED"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 2
        actions = [
            "1. Contact leaders of units missing from web (see 'Key Three Only' sheet)",
            "2. Verify registration status of web-only units (see 'Web Only' sheet)",
            "3. Review data quality issues for all units (see 'Data Quality' sheet)",
            "4. Coordinate with council registrar for database updates"
        ]
        
        for action in actions:
            ws[f'A{row}'] = action
            row += 1
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 30
    
    def _create_key_three_only_sheet(self, workbook):
        """Create sheet for units missing from web"""
        ws = workbook.create_sheet("Key Three Only")
        
        # Header
        ws['A1'] = "Units Missing from Web - Immediate Action Required"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Get Key Three only units
        key_three_only_units = [
            result for result in self.validation_data['validation_results']
            if result['status'] == 'key_three_only'
        ]
        
        if not key_three_only_units:
            ws['A3'] = "No units missing from web - Excellent!"
            return
        
        # Create data for table
        data = []
        for result in key_three_only_units:
            key_three_data = result.get('key_three_data', {})
            
            data.append({
                'Unit': result['unit_key'],
                'District': key_three_data.get('district', 'Unknown'),
                'Town': key_three_data.get('unit_town', 'Unknown'),
                'Chartered Organization': key_three_data.get('chartered_org_name', 'Unknown'),
                'Committee Chair': key_three_data.get('committee_chair_name', 'N/A'),
                'Chair Email': key_three_data.get('committee_chair_email', 'N/A'),
                'Cubmaster/Scoutmaster': key_three_data.get('unit_leader_name', 'N/A'),
                'Leader Email': key_three_data.get('unit_leader_email', 'N/A'),
                'Action Required': 'Contact leaders about establishing web presence'
            })
        
        # Create DataFrame and write to sheet
        df = pd.DataFrame(data)
        
        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Write data
        for row_num, row_data in enumerate(df.values, 4):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_web_only_sheet(self, workbook):
        """Create sheet for units on web but not in Key Three"""
        ws = workbook.create_sheet("Web Only")
        
        # Header
        ws['A1'] = "Units on Web but Not in Key Three Database"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Get web only units
        web_only_units = [
            result for result in self.validation_data['validation_results']
            if result['status'] == 'web_only'
        ]
        
        if not web_only_units:
            ws['A3'] = "No unauthorized web units found - Good data integrity!"
            return
        
        # Create data for table
        data = []
        for result in web_only_units:
            scraped_data = result.get('scraped_data', {})
            
            data.append({
                'Unit': result['unit_key'],
                'District': scraped_data.get('district', 'Unknown'),
                'Town': scraped_data.get('unit_town', 'Unknown'),
                'Chartered Organization': scraped_data.get('chartered_org_name', 'Unknown'),
                'Meeting Location': scraped_data.get('meeting_location', 'N/A'),
                'Meeting Schedule': f"{scraped_data.get('meeting_day', 'N/A')} {scraped_data.get('meeting_time', '')}".strip(),
                'Contact Email': scraped_data.get('contact_email', 'N/A'),
                'Website': scraped_data.get('website', 'N/A'),
                'Action Required': 'Verify unit registration status with registrar'
            })
        
        # Create DataFrame and write to sheet
        df = pd.DataFrame(data)
        
        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        # Write data
        for row_num, row_data in enumerate(df.values, 4):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_data_quality_sheet(self, workbook):
        """Create sheet for units with data quality issues"""
        ws = workbook.create_sheet("Data Quality Issues")
        
        # Header
        ws['A1'] = "Units with Data Quality Issues"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Get units with issues
        units_with_issues = [
            result for result in self.validation_data['validation_results']
            if result.get('issues') and result['status'] == 'both_sources'
        ]
        
        if not units_with_issues:
            ws['A3'] = "No data quality issues found - Excellent!"
            return
        
        # Create data for table
        data = []
        for result in units_with_issues:
            scraped_data = result.get('scraped_data', {})
            key_three_data = result.get('key_three_data', {})
            
            issues_text = "; ".join(result.get('issues', []))
            
            data.append({
                'Unit': result['unit_key'],
                'District': scraped_data.get('district', 'Unknown'),
                'Town': scraped_data.get('unit_town', 'Unknown'),
                'Issues': issues_text,
                'Meeting Location': scraped_data.get('meeting_location', 'Missing'),
                'Meeting Day': scraped_data.get('meeting_day', 'Missing'),
                'Meeting Time': scraped_data.get('meeting_time', 'Missing'),
                'Contact Email': scraped_data.get('contact_email', 'Missing'),
                'Unit Leader': key_three_data.get('unit_leader_name', 'N/A'),
                'Leader Email': key_three_data.get('unit_leader_email', 'N/A')
            })
        
        # Sort by district and unit
        data.sort(key=lambda x: (x['District'], x['Unit']))
        
        # Create DataFrame and write to sheet
        df = pd.DataFrame(data)
        
        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Write data
        for row_num, row_data in enumerate(df.values, 4):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_complete_unit_listing_sheet(self, workbook):
        """Create complete unit listing with validation status"""
        ws = workbook.create_sheet("Complete Unit Listing")
        
        # Header
        ws['A1'] = "Complete Heart of New England Council Unit Listing"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Create data for all units
        data = []
        for result in self.validation_data['validation_results']:
            scraped_data = result.get('scraped_data', {})
            key_three_data = result.get('key_three_data', {})
            
            # Determine primary data source
            if result['status'] == 'both_sources':
                district = scraped_data.get('district', key_three_data.get('district', 'Unknown'))
                town = scraped_data.get('unit_town', key_three_data.get('unit_town', 'Unknown'))
                chartered_org = scraped_data.get('chartered_org_name', key_three_data.get('chartered_org_name', 'Unknown'))
            elif result['status'] == 'key_three_only':
                district = key_three_data.get('district', 'Unknown')
                town = key_three_data.get('unit_town', 'Unknown')
                chartered_org = key_three_data.get('chartered_org_name', 'Unknown')
            else:  # web_only
                district = scraped_data.get('district', 'Unknown')
                town = scraped_data.get('unit_town', 'Unknown')
                chartered_org = scraped_data.get('chartered_org_name', 'Unknown')
            
            status_display = {
                'both_sources': 'Both Sources ✅',
                'key_three_only': 'Key Three Only ⚠️',
                'web_only': 'Web Only ❌'
            }.get(result['status'], result['status'])
            
            data.append({
                'Unit': result['unit_key'],
                'District': district,
                'Town': town,
                'Chartered Organization': chartered_org,
                'Validation Status': status_display,
                'Web Presence': 'Yes' if scraped_data else 'No',
                'Key Three Registration': 'Yes' if key_three_data else 'No',
                'Issues Count': len(result.get('issues', []))
            })
        
        # Sort by district, then unit type, then unit number
        data.sort(key=lambda x: (x['District'], x['Unit']))
        
        # Create DataFrame and write to sheet
        df = pd.DataFrame(data)
        
        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Write data with conditional formatting
        for row_num, row_data in enumerate(df.values, 4):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Color code by status
                if col_num == 5:  # Validation Status column
                    if 'Both Sources' in str(cell_value):
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif 'Key Three Only' in str(cell_value):
                        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                    elif 'Web Only' in str(cell_value):
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

def main():
    """Generate commissioner data quality report"""
    
    print("📋 Generating Commissioner Data Quality Report")
    
    generator = CommissionerReportGenerator()
    
    # Load validation data
    if not generator.load_validation_data():
        return
    
    # Generate report
    report_path = generator.create_commissioner_report()
    
    # Get actual numbers for display
    summary = generator.validation_data['validation_summary']
    key_three_only = summary['status_breakdown']['key_three_only']
    web_only = summary['status_breakdown']['web_only']
    issues = summary['units_with_issues']
    
    print(f"\n✅ Commissioner Report Generated Successfully!")
    print(f"📁 Report saved to: {report_path}")
    print(f"\n📊 Report Contents:")
    print(f"   • Executive Summary with key metrics")
    print(f"   • Units missing from web ({key_three_only} units requiring contact)")
    print(f"   • Units on web only ({web_only} units requiring verification)")
    print(f"   • Data quality issues ({issues} units needing updates)")
    print(f"   • Complete unit listing with validation status")
    
    return report_path

if __name__ == "__main__":
    main()