#!/usr/bin/env python3
"""
BeAScout Quality Report Generator
Creates comprehensive Excel reports with quality scores, grades, and recommendations
Organized by district sheets with detailed unit information for commissioners
"""

import json
import re
import sys
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any

# Add project root to path for imports
sys.path.append(str(Path(__file__).parent.parent.parent))

try:
    from src.pipeline.analysis.quality_scorer import UnitQualityScorer
except ImportError:
    # Fallback for when running from different directory
    import os
    current_dir = Path(__file__).parent
    quality_scorer_path = current_dir.parent / "analysis" / "quality_scorer.py"
    import importlib.util
    spec = importlib.util.spec_from_file_location("quality_scorer", quality_scorer_path)
    quality_scorer_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(quality_scorer_module)
    UnitQualityScorer = quality_scorer_module.UnitQualityScorer

class BeAScoutQualityReportGenerator:
    """
    Generates comprehensive BeAScout Quality Reports organized by district
    Includes quality scores, grades, recommendations, and Key Three contact information
    """
    
    def __init__(self):
        self.quality_data = None
        self.key_three_data = None
        self.scorer = UnitQualityScorer()
        
    def load_quality_data(self, quality_file: str = 'data/raw/all_units_comprehensive_scored.json',
                          validation_file: str = 'data/output/enhanced_three_way_validation_results.json') -> bool:
        """Load quality-scored unit data, Key Three information, and missing units"""
        try:
            # Load quality-scored units
            with open(quality_file, 'r', encoding='utf-8') as f:
                self.quality_data = json.load(f)
                print(f"📊 Loaded quality data: {len(self.quality_data['units_with_scores'])} units")
            
            # Load validation data to get Key Three info and missing units
            with open(validation_file, 'r', encoding='utf-8') as f:
                validation_data = json.load(f)
                
                # Create lookup for Key Three member information
                self.key_three_data = {}
                missing_units = []
                
                for result in validation_data['validation_results']:
                    if 'key_three_data' in result:
                        unit_key = result['unit_key']
                        self.key_three_data[unit_key] = result['key_three_data']
                        
                        # Collect Key Three-only units (missing from BeAScout)
                        if result['status'] == 'key_three_only':
                            missing_units.append(self._create_missing_unit_record(result['key_three_data']))
                
                # Add missing units to quality data with score=0, grade="N/A"
                self.quality_data['units_with_scores'].extend(missing_units)
                self.quality_data['total_units'] = len(self.quality_data['units_with_scores'])
                
                print(f"📊 Loaded Key Three data for {len(self.key_three_data)} units")
                print(f"📊 Added {len(missing_units)} units missing from BeAScout")
                
                # Store validation summary for executive summary
                self.validation_summary = validation_data['validation_summary']
            
            return True
        except Exception as e:
            print(f"❌ Failed to load data: {e}")
            return False
    
    def _create_missing_unit_record(self, key_three_data: Dict) -> Dict:
        """Create a unit record for Key Three-only units with score=0, grade=N/A"""
        return {
            'unit_key': key_three_data.get('unit_key', ''),
            'unit_type': key_three_data.get('unit_type', ''),
            'unit_number': key_three_data.get('unit_number', ''),
            'unit_town': key_three_data.get('unit_town', ''),
            'chartered_organization': key_three_data.get('chartered_organization', ''),
            'district': key_three_data.get('district', ''),
            'data_source': 'key_three_only',
            # Empty contact/meeting information since not present in BeAScout
            'meeting_location': '',
            'meeting_day': '',
            'meeting_time': '',
            'contact_email': '',
            'contact_person': '',
            'phone_number': '',
            'website': '',
            'description': '',
            'specialty': '',
            # Quality scoring for missing units
            'completeness_score': 0,
            'completeness_grade': 'N/A',
            'recommendations': [],
            # Mark as missing from BeAScout
            'missing_from_beascout': True
        }
    
    def create_quality_report(self, output_path: str = None) -> str:
        """
        Create comprehensive BeAScout Quality Report organized by district
        
        Returns:
            Path to generated Excel file
        """
        if not self.quality_data:
            raise ValueError("No quality data loaded")
        
        # Generate timestamped filename if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"data/output/reports/BeAScout_Quality_Report_{timestamp}.xlsx"
        
        # Ensure output directory exists
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        
        # Create Executive Summary first
        self._create_executive_summary_sheet(workbook)
        
        # Create district-specific sheets
        self._create_district_sheets(workbook)
        
        # Save workbook
        workbook.save(output_path)
        print(f"\n📋 BeAScout Quality Report saved: {output_path}")
        
        return output_path
    
    def _create_executive_summary_sheet(self, workbook):
        """Create executive summary sheet with quality metrics and legend"""
        ws = workbook.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "BeAScout Quality Report"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Council name and details
        ws['A2'] = "Heart of New England Council"
        ws['A2'].font = Font(size=14, bold=True)
        
        # Report generation info
        ws['A3'] = f"Generation Date/Time: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A3'].font = Font(size=12)
        
        # Data sources - include Key Three spreadsheet name
        key_three_filename = "Key 3 08-22-2025.xlsx"  # TODO: Extract from actual filename
        ws['A4'] = f"Data Sources: BeAScout.org (10-mile radius) + JoinExploring.org (20-mile) + {key_three_filename}"
        ws['A4'].font = Font(size=10)
        
        ws['A5'] = f"Last Complete BeAScout Data Retrieval: {self.quality_data.get('extraction_timestamp', 'Unknown')[:10]}"
        ws['A5'].font = Font(size=10)
        
        # Quality metrics - remove extra spacing
        row = 7
        ws[f'A{row}'] = "QUALITY OVERVIEW"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        # Calculate metrics from quality data including missing units
        total_units = self.quality_data['total_units']
        
        # Calculate average score only for units with BeAScout data
        beascout_units = [u for u in self.quality_data['units_with_scores'] if not u.get('missing_from_beascout', False)]
        if beascout_units:
            avg_score = sum(u.get('completeness_score', 0) for u in beascout_units) / len(beascout_units)
        else:
            avg_score = 0
        
        # Count missing units
        missing_units_count = len([u for u in self.quality_data['units_with_scores'] if u.get('missing_from_beascout', False)])
        
        # Count units by district
        district_counts = {}
        town_counts = set()
        grade_counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0, 'N/A': 0}
        
        for unit in self.quality_data['units_with_scores']:
            district = unit.get('district', 'Unknown')
            district_counts[district] = district_counts.get(district, 0) + 1
            town_counts.add(unit.get('unit_town', 'Unknown'))
            grade = unit.get('completeness_grade', 'F')
            if grade in grade_counts:
                grade_counts[grade] += 1
        
        row += 1  # Reduced spacing
        metrics = [
            ("Total Units Analyzed", total_units),
            ("Units missing from BeAScout", missing_units_count),
            ("Units across Towns", f"{total_units} units across {len(town_counts)} towns"),
            ("Average Quality Score", f"{avg_score:.1f}%"),
            ("Quality Grade Distribution:", ""),
            ("  • Grade A (90%+)", f"{grade_counts['A']} units ({grade_counts['A']/total_units*100:.1f}%)"),
            ("  • Grade B (80-89%)", f"{grade_counts['B']} units ({grade_counts['B']/total_units*100:.1f}%)"),
            ("  • Grade C (70-79%)", f"{grade_counts['C']} units ({grade_counts['C']/total_units*100:.1f}%)"),
            ("  • Grade D (60-69%)", f"{grade_counts['D']} units ({grade_counts['D']/total_units*100:.1f}%)"),
            ("  • Grade F (<60%)", f"{grade_counts['F']} units ({grade_counts['F']/total_units*100:.1f}%)"),
            ("  • Grade N/A (Missing)", f"{grade_counts['N/A']} units ({grade_counts['N/A']/total_units*100:.1f}%)"),
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = str(value)
            if not metric.startswith('  •'):
                ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # District breakdown - reduced spacing
        row += 1  # Reduced spacing
        ws[f'A{row}'] = "UNITS BY DISTRICT"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 1  # Reduced spacing
        for district, count in sorted(district_counts.items()):
            ws[f'A{row}'] = district
            ws[f'B{row}'] = f"{count} units"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Quality Issue Legend with restructured format - reduced spacing
        row += 1  # Reduced spacing
        ws[f'A{row}'] = "QUALITY ISSUE LEGEND"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 1  # Reduced spacing
        # Required Information subsection
        ws[f'A{row}'] = "Required Information:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        required_items = [
            ("Missing meeting location", "Unit needs a physical meeting location with street address for parents and youth to find meetings"),
            ("Missing meeting day", "Unit needs to specify which day(s) of the week meetings are held"),
            ("Missing meeting time", "Unit needs to specify what time meetings start and end"),
            ("Missing contact email", "Unit needs a contact email address for inquiries and communication"),
        ]
        for issue, explanation in required_items:
            ws[f'A{row}'] = f"  • {issue}"
            ws[f'B{row}'] = explanation
            row += 1
        
        row += 1
        # Needs Improvement subsection  
        ws[f'A{row}'] = "Needs Improvement:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        quality_items = [
            ("PO Box location", "Replace PO Box with physical meeting location so parents and youth can find meetings"),
            ("Personal email", "Use unit-specific email instead of personal email for better continuity and professionalism"),
        ]
        for issue, explanation in quality_items:
            ws[f'A{row}'] = f"  • {issue}"
            ws[f'B{row}'] = explanation
            row += 1
            
        row += 1
        # Recommended Information subsection
        ws[f'A{row}'] = "Recommended Information:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        recommended_items = [
            ("Missing contact person", "Adding a contact person name helps parents and youth know who to reach out to"),
            ("Missing phone number", "Phone contact provides immediate communication option for urgent questions"),
            ("Missing website", "Unit-specific website increases visibility and provides additional information for families"),
            ("Missing description", "Informative description helps attract new members by explaining unit activities and culture"),
        ]
        for issue, explanation in recommended_items:
            ws[f'A{row}'] = f"  • {issue}"
            ws[f'B{row}'] = explanation
            row += 1
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 60
    
    def _create_district_sheets(self, workbook):
        """Create separate sheets for each district with detailed unit information"""
        # Group units by district
        units_by_district = {}
        for unit in self.quality_data['units_with_scores']:
            district = unit.get('district', 'Unknown')
            if district not in units_by_district:
                units_by_district[district] = []
            units_by_district[district].append(unit)
        
        # Create sheet for each district
        for district_name in sorted(units_by_district.keys()):
            if district_name != 'Unknown':  # Skip unknown districts
                self._create_district_sheet(workbook, district_name, units_by_district[district_name])
    
    def _create_district_sheet(self, workbook, district_name: str, units: List[Dict]):
        """Create individual district sheet with detailed unit information"""
        ws = workbook.create_sheet(district_name)
        
        # Header section
        ws['A1'] = "BeAScout Quality Report"
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A2'] = "Heart of New England Council"
        ws['A2'].font = Font(size=14, bold=True)
        
        ws['A3'] = f"{district_name} District"
        ws['A3'].font = Font(size=14, bold=True)
        
        ws['A4'] = f"Generation Date/Time: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A4'].font = Font(size=10)
        
        key_three_filename = "Key 3 08-22-2025.xlsx"  # TODO: Extract from actual filename
        ws['A5'] = f"Data Sources: BeAScout.org (10-mile radius) + JoinExploring.org (20-mile) + {key_three_filename}"
        ws['A5'].font = Font(size=10)
        
        ws['A6'] = f"Last Complete BeAScout Data Retrieval: {self.quality_data.get('extraction_timestamp', 'Unknown')[:10]}"
        ws['A6'].font = Font(size=10)
        
        # Count units and towns for this district
        towns = set(unit.get('unit_town', 'Unknown') for unit in units)
        ws['A7'] = f"{len(units)} Units across {len(towns)} towns"
        ws['A7'].font = Font(size=12, bold=True)
        
        # Create column headers (row 9) with reordered contact columns
        headers = [
            "Unit Identifier", "Chartered Organization", "Town", "Quality Score", "Quality Grade",
            "Missing Information", "Quality Issues", "Recommended Improvements", "Meeting Location",
            "Meeting Day", "Meeting Time", "Contact Person", "Contact Email", "Contact Phone Number",
            "Unit Website", "Key Three - First Person", "Key Three - Second Person", "Key Three - Third Person"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Sort units by quality grade (worst first), then score, then unit identifier
        # Grade hierarchy: N/A (worst) < F < D < C < B < A (best)
        def sort_key(unit):
            grade = unit.get('completeness_grade', 'F')
            score = unit.get('completeness_score', 0)
            unit_key = unit.get('unit_key', '')
            
            # Map grades to sort order (lower numbers = worse grades = higher priority)
            grade_order = {'N/A': 0, 'F': 1, 'D': 2, 'C': 3, 'B': 4, 'A': 5}
            grade_priority = grade_order.get(grade, 1)  # Default to F priority
            
            return (grade_priority, score, unit_key)
        
        sorted_units = sorted(units, key=sort_key)
        
        # Add unit data starting from row 10
        for row_num, unit in enumerate(sorted_units, 10):
            self._populate_unit_row(ws, row_num, unit)
        
        # Auto-adjust column widths
        column_widths = [20, 30, 15, 12, 12, 25, 25, 25, 25, 12, 12, 20, 18, 18, 20, 35, 35, 35]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Freeze columns A and B (and header rows)
        ws.freeze_panes = 'C10'
        
        # Apply professional formatting to all data cells
        self._apply_district_sheet_formatting(ws, len(sorted_units), len(headers))
    
    def _populate_unit_row(self, ws, row_num: int, unit: Dict):
        """Populate a single unit row with all required information"""
        # Get Key Three information if available
        unit_key = unit.get('unit_key', '')
        key_three_info = self.key_three_data.get(unit_key, {})
        key_three_members = key_three_info.get('key_three_members', [])
        
        # Categorize recommendations
        recommendations = unit.get('recommendations', [])
        missing_info = [rec for rec in recommendations if rec.startswith('REQUIRED_')]
        quality_issues = [rec for rec in recommendations if rec.startswith('QUALITY_')]
        recommended_improvements = [rec for rec in recommendations if rec.startswith('RECOMMENDED_') or rec.startswith('CONTENT_')]
        
        # Get readable descriptions
        missing_descriptions = self.scorer.get_recommendation_descriptions(missing_info)
        quality_descriptions = self.scorer.get_recommendation_descriptions(quality_issues)
        recommended_descriptions = self.scorer.get_recommendation_descriptions(recommended_improvements)
        
        # Format Key Three member information (remove labels, multi-line, no unhelpful N/A)
        def format_key_three_member(member_info):
            if not member_info or member_info.get('fullname') == 'None':
                return "None"
            
            # Only include actual data, skip N/A values
            lines = []
            position = member_info.get('position', '')
            name = member_info.get('fullname', '')
            email = member_info.get('email', '')
            phone = member_info.get('phone', '')
            
            # Only add lines that have meaningful content
            if position and position != 'N/A':
                lines.append(position)
            if name and name != 'N/A':
                lines.append(name)
            # Skip address since not available in Key Three data
            if email and email != 'N/A':
                lines.append(email)
            if phone and phone != 'N/A':
                lines.append(phone)
            
            return '\n'.join(lines) if lines else 'None'
        
        key_three_1 = format_key_three_member(key_three_members[0] if len(key_three_members) > 0 else {})
        key_three_2 = format_key_three_member(key_three_members[1] if len(key_three_members) > 1 else {})
        key_three_3 = format_key_three_member(key_three_members[2] if len(key_three_members) > 2 else {})
        
        # Format meeting location with multi-line (facility\nstreet address\ncity/state/zip)
        meeting_location = unit.get('meeting_location', '')
        if meeting_location:
            # Smart parsing to avoid breaking multi-word town names
            # Only break on commas that aren't part of town names
            parts = []
            current_part = ''
            
            # Split on commas but be careful with multi-word towns
            for part in meeting_location.split(', '):
                part = part.strip()
                if part:
                    # Check if this looks like a state/zip (e.g., "MA 01234")
                    if re.match(r'^[A-Z]{2}\s+\d{5}', part):
                        # This is likely state/zip, add to parts
                        parts.append(part)
                    elif len(parts) == 0:
                        # First part is likely facility/building
                        parts.append(part)
                    elif re.match(r'^\d+\s+', part):
                        # Starts with number, likely street address
                        parts.append(part)
                    else:
                        # Could be continuation of address or town name
                        # If previous part looks incomplete, combine
                        if parts and not re.match(r'^\d+\s+.*\s+(Street|St|Road|Rd|Avenue|Ave|Lane|Ln|Drive|Dr|Way|Circle|Cir)', parts[-1], re.IGNORECASE):
                            parts[-1] = parts[-1] + ', ' + part
                        else:
                            parts.append(part)
            
            formatted_location = '\n'.join(parts) if parts else meeting_location
        else:
            formatted_location = ''
        
        # Populate row data with reordered contact columns and multi-line formatting
        row_data = [
            unit_key,
            unit.get('chartered_organization', ''),
            unit.get('unit_town', ''),
            unit.get('completeness_score', 0),
            unit.get('completeness_grade', 'F'),
            '\n'.join(missing_descriptions),  # Each issue on separate line
            '\n'.join(quality_descriptions),  # Each issue on separate line
            '\n'.join(recommended_descriptions),  # Each issue on separate line
            formatted_location,  # Multi-line meeting location
            unit.get('meeting_day', ''),
            unit.get('meeting_time', ''),
            unit.get('contact_person', ''),  # Reordered: Contact Person first
            unit.get('contact_email', ''),   # Then Contact Email
            unit.get('phone_number', ''),    # Then Contact Phone
            unit.get('website', ''),
            key_three_1,
            key_three_2,
            key_three_3
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            # Handle 0 values explicitly (don't convert to empty string)
            if value is None or value == '':
                cell.value = ""
            elif col_num == 4:  # Quality Score column - use numeric format, left-aligned
                cell.value = float(value) if isinstance(value, (int, float)) else value
            else:
                cell.value = str(value)
            
            # Set alignment - left-aligned for Quality Score, default for others
            if col_num == 4:
                cell.alignment = Alignment(horizontal='left', wrap_text=True, vertical='top')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Color code quality grades including N/A for missing units
            if col_num == 5:  # Quality Grade column
                grade = str(value)
                if grade == 'A':
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif grade == 'B':
                    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                elif grade == 'C':
                    cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                elif grade == 'D':
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                elif grade == 'F':
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                elif grade == 'N/A':
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for missing units
    
    def _apply_district_sheet_formatting(self, ws, num_data_rows: int, num_columns: int):
        """Apply professional formatting to district sheet with borders"""
        from openpyxl.styles import Border, Side
        
        # Define border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply formatting to header row (row 9)
        for col_num in range(1, num_columns + 1):
            cell = ws.cell(row=9, column=col_num)
            cell.border = thin_border
            
        # Apply formatting to data rows (starting from row 10)
        for row_num in range(10, 10 + num_data_rows):
            for col_num in range(1, num_columns + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border

def main():
    """Generate BeAScout Quality Report organized by districts"""
    
    print("📋 Generating BeAScout Quality Report")
    
    generator = BeAScoutQualityReportGenerator()
    
    # Load quality and Key Three data
    if not generator.load_quality_data():
        return
    
    # Generate report
    report_path = generator.create_quality_report()
    
    # Get statistics for display
    total_units = generator.quality_data['total_units']
    avg_score = generator.quality_data.get('average_completeness_score', 0)
    
    # Count districts
    districts = set()
    for unit in generator.quality_data['units_with_scores']:
        district = unit.get('district', 'Unknown')
        if district != 'Unknown':
            districts.add(district)
    
    print(f"\n✅ BeAScout Quality Report Generated Successfully!")
    print(f"📁 Report saved to: {report_path}")
    print(f"\n📊 Report Contents:")
    print(f"   • Executive Summary with quality metrics and legend")
    print(f"   • {len(districts)} District sheets with detailed unit information")
    print(f"   • {total_units} total units with quality scores (avg: {avg_score}%)")
    print(f"   • 18-column format with quality grades, recommendations, and Key Three contacts")
    
    return report_path

if __name__ == "__main__":
    main()