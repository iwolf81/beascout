#!/usr/bin/env python3
"""
BeAScout Quality Report Generator
Creates comprehensive Excel reports with quality scores, grades, and recommendations
Organized by district sheets with detailed unit information for commissioners
"""

import json
import re
import sys
import argparse
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any

# Add project root to path for imports
project_root = Path(__file__).parent.parent.parent.parent
sys.path.append(str(project_root))

from src.pipeline.core.session_utils import SessionManager, session_logging

def load_excluded_units():
    """Load list of units to exclude from all reports and emails"""
    excluded_file = project_root / "data/config/excluded_units.json"
    if not excluded_file.exists():
        return set()

    try:
        with open(excluded_file, 'r') as f:
            data = json.load(f)
        return {unit['unit_key'] for unit in data.get('excluded_units', [])}
    except Exception as e:
        print(f"⚠️  Warning: Could not load excluded units: {e}")
        return set()

class ReportColumns:
    """Centralized column definitions for BeAScout Quality Reports"""
    
    # Column numbers (1-indexed for Excel)
    UNIT_IDENTIFIER = 1
    CHARTERED_ORG = 2
    TOWN = 3
    ZIP_CODE = 4
    QUALITY_SCORE = 5
    QUALITY_GRADE = 6
    MISSING_INFO = 7
    QUALITY_ISSUES = 8
    RECOMMENDED_IMPROVEMENTS = 9
    MEETING_LOCATION = 10
    MEETING_DAY = 11
    MEETING_TIME = 12
    SPECIALTY = 13
    CONTACT_EMAIL = 14
    CONTACT_PERSON = 15
    CONTACT_PHONE = 16
    UNIT_WEBSITE = 17
    KEY_THREE_1 = 18
    KEY_THREE_2 = 19
    KEY_THREE_3 = 20
    
    # Total number of columns
    TOTAL_COLUMNS = 20
    
    # Column headers in order
    HEADERS = [
        "Unit Identifier", "Chartered Organization", "Town", "Zip Code", "Quality Score", "Quality Grade",
        "Missing Information", "Quality Issues", "Recommended Improvements", "Meeting Location",
        "Meeting Day", "Meeting Time", "Specialty (Crews only)", "Contact Email", "Contact Person", "Contact Phone Number",
        "Unit Website", "Key Three (1)", "Key Three (2)", "Key Three (3)"
    ]
    
    # Column widths (in Excel units)
    WIDTHS = [20, 30, 15, 10, 12, 12, 25, 25, 25, 25, 12, 12, 15, 20, 18, 18, 20, 35, 35, 35]
    
    # Column categories for styling
    IDENTITY_COLUMNS = [UNIT_IDENTIFIER, CHARTERED_ORG, TOWN, ZIP_CODE]
    QUALITY_COLUMNS = [QUALITY_SCORE, QUALITY_GRADE, MISSING_INFO, QUALITY_ISSUES, RECOMMENDED_IMPROVEMENTS]
    MEETING_COLUMNS = [MEETING_LOCATION, MEETING_DAY, MEETING_TIME, SPECIALTY, CONTACT_EMAIL, CONTACT_PERSON, CONTACT_PHONE, UNIT_WEBSITE]
    KEY_THREE_COLUMNS = [KEY_THREE_1, KEY_THREE_2, KEY_THREE_3]
    
    # Special formatting columns
    CLICKABLE_EMAIL_COLUMNS = [CONTACT_EMAIL] + KEY_THREE_COLUMNS
    CLICKABLE_WEBSITE_COLUMNS = [UNIT_WEBSITE]
    NUMERIC_COLUMNS = [QUALITY_SCORE]
    
    @classmethod
    def get_column_number(cls, column_name: str) -> int:
        """Get column number by name (for dynamic lookups)"""
        return getattr(cls, column_name.upper())
    
    @classmethod
    def is_in_category(cls, col_num: int, category: str) -> bool:
        """Check if column is in a specific category"""
        category_map = {
            'identity': cls.IDENTITY_COLUMNS,
            'quality': cls.QUALITY_COLUMNS,
            'meeting': cls.MEETING_COLUMNS,
            'key_three': cls.KEY_THREE_COLUMNS
        }
        return col_num in category_map.get(category, [])

# Add project root to path for imports
sys.path.append(str(Path(__file__).parent.parent.parent))

try:
    from src.pipeline.core.quality_scorer import UnitQualityScorer
    from src.pipeline.core.unit_identifier import UnitIdentifierNormalizer
except ImportError:
    # Fallback for when running from different directory
    import os
    current_dir = Path(__file__).parent
    import importlib.util
    
    # Load quality scorer
    quality_scorer_path = current_dir.parent / "core" / "quality_scorer.py"
    spec = importlib.util.spec_from_file_location("quality_scorer", quality_scorer_path)
    quality_scorer_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(quality_scorer_module)
    UnitQualityScorer = quality_scorer_module.UnitQualityScorer
    
    # Load unit identifier normalizer
    unit_identifier_path = current_dir.parent / "core" / "unit_identifier.py"
    spec = importlib.util.spec_from_file_location("unit_identifier", unit_identifier_path)
    unit_identifier_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(unit_identifier_module)
    UnitIdentifierNormalizer = unit_identifier_module.UnitIdentifierNormalizer

def load_town_zip_mapping():
    """Load town to zip code mapping for HNE Council"""
    # Try relative to current working directory first
    zip_file_path = Path("data/zipcodes/hne_council_zipcodes.json")
    if not zip_file_path.exists():
        # Try relative to script location
        zip_file_path = Path(__file__).parent.parent.parent / "data" / "zipcodes" / "hne_council_zipcodes.json"
    
    try:
        with open(zip_file_path, 'r') as f:
            zip_data = json.load(f)
            return zip_data.get('town_zipcodes', {})
    except FileNotFoundError:
        print(f"⚠️  Zip code mapping file not found: {zip_file_path}")
        return {}

def get_zip_code_for_town(town_name: str, town_zip_mapping: dict) -> str:
    """Get primary zip code for a town name, including village-to-town mappings"""
    if not town_name or not town_zip_mapping:
        return ""
    
    # Clean up town name variations
    clean_town = town_name.strip()
    
    # Village-to-town mappings for HNE Council
    village_to_town = {
        "Fiskdale": "Sturbridge",  # Fiskdale is a village within Sturbridge
        "Whitinsville": "Northbridge",  # Whitinsville is a village within Northbridge
    }
    
    # Check if this is a known village
    if clean_town in village_to_town:
        parent_town = village_to_town[clean_town]
        if parent_town in town_zip_mapping:
            zip_codes = town_zip_mapping[parent_town]
            return zip_codes[0] if zip_codes else ""
    
    # Direct lookup
    if clean_town in town_zip_mapping:
        zip_codes = town_zip_mapping[clean_town]
        return zip_codes[0] if zip_codes else ""  # Return first zip code
    
    # Try variations for compound town names
    if clean_town.startswith("E. "):
        base_town = clean_town[3:]  # Remove "E. " prefix
        if base_town in town_zip_mapping:
            zip_codes = town_zip_mapping[base_town]
            return zip_codes[0] if zip_codes else ""
    
    # Try "North " prefix variations
    if clean_town.startswith("North "):
        base_town = clean_town[6:]  # Remove "North " prefix  
        if base_town in town_zip_mapping:
            zip_codes = town_zip_mapping[base_town]
            return zip_codes[0] if zip_codes else ""
    
    return ""  # No mapping found

class BeAScoutQualityReportGenerator:
    """
    Generates comprehensive BeAScout Quality Reports organized by district
    Includes quality scores, grades, recommendations, and Key Three contact information
    """
    
    def __init__(self, key_three_filename: str = None, scraped_session_id: str = None):
        self.quality_data = None
        self.key_three_data = None
        self.scorer = UnitQualityScorer()
        self.town_zip_mapping = load_town_zip_mapping()
        self.key_three_filename = key_three_filename or "[Key Three filename not specified]"
        self.scraped_session_id = scraped_session_id

    def _format_scraped_timestamp(self) -> str:
        """Format scraped session timestamp for display"""
        if not self.scraped_session_id:
            return ""

        try:
            from datetime import datetime
            # Parse timestamp: YYYYMMDD_HHMMSS
            date_part = self.scraped_session_id[:8]
            time_part = self.scraped_session_id[9:]

            date_obj = datetime.strptime(date_part, "%Y%m%d")
            time_obj = datetime.strptime(time_part, "%H%M%S")

            formatted_date = date_obj.strftime("%B %d, %Y")
            formatted_time = time_obj.strftime("%I:%M %p")

            return f" - Data scraped: {formatted_date} at {formatted_time}"
        except Exception:
            return f" - Scraped session: {self.scraped_session_id}"
        
    def load_quality_data(self, quality_file: str = 'data/raw/all_units_comprehensive_scored.json',
                          validation_file: str = 'data/output/enhanced_three_way_validation_results.json') -> bool:
        """Load unit data with quality tags already calculated during HTML parsing"""
        try:
            # Load excluded units
            excluded_units = load_excluded_units()
            if excluded_units:
                print(f"🚫 Loaded {len(excluded_units)} excluded units from config")

            # Load unit data with integrated quality scoring
            with open(quality_file, 'r', encoding='utf-8') as f:
                raw_data = json.load(f)

                # Handle both data formats: scraped_units or units_with_scores
                all_units = raw_data.get('scraped_units', raw_data.get('units_with_scores', []))

                # Filter out excluded units
                units = [u for u in all_units if u.get('unit_key', '') not in excluded_units]
                excluded_count = len(all_units) - len(units)

                print(f"📊 Loaded {len(units)} units with quality data")
                if excluded_count > 0:
                    print(f"🚫 Excluded {excluded_count} units from report")
            
            # Load scraping timestamp from the comprehensive data source tracking
            scraping_timestamp = ''
            scraped_data_source = ''
            try:
                # Check if comprehensive data has source tracking
                if raw_data.get('session_summary', {}).get('start_time'):
                    scraping_timestamp = raw_data['session_summary']['start_time']
                    scraped_data_source = raw_data.get('scraped_data_source', '')
                    print(f"📅 Found scraping timestamp from source tracking: {scraping_timestamp}")
                    if scraped_data_source:
                        print(f"📁 Source directory: {scraped_data_source}")
                else:
                    print(f"⚠️ No source tracking found in comprehensive data")
                    scraping_timestamp = ''
            except Exception as e:
                print(f"⚠️ Could not load scraping timestamp from source tracking: {e}")
                scraping_timestamp = ''
            
            # Quality data is now integrated - no need for separate scoring
            self.quality_data = {
                'total_units': len(units),
                'units_with_scores': units,
                'average_score': 0.0,
                'scoring_summary': {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0},
                'extraction_timestamp': raw_data.get('extraction_timestamp', ''),  # Processing timestamp
                'scraping_timestamp': scraping_timestamp  # Original scraping timestamp from session_summary.json
            }
            
            # Calculate summary statistics from integrated quality data
            total_score = 0.0
            for unit in self.quality_data['units_with_scores']:
                grade = unit.get('completeness_grade', 'F')
                score = unit.get('completeness_score', 0.0)
                self.quality_data['scoring_summary'][grade] += 1
                total_score += score
            
            if len(self.quality_data['units_with_scores']) > 0:
                self.quality_data['average_score'] = round(total_score / len(self.quality_data['units_with_scores']), 1)
            
            print(f"📊 Quality data integrated: avg score {self.quality_data['average_score']}%")
            
            # Load validation data to get Key Three info and missing units
            with open(validation_file, 'r', encoding='utf-8') as f:
                validation_data = json.load(f)
                
                # Create lookup for Key Three member information
                self.key_three_data = {}
                missing_units = []
                
                for result in validation_data['validation_results']:
                    if 'key_three_data' in result:
                        unit_key = result['unit_key']

                        # Skip excluded units
                        if unit_key in excluded_units:
                            continue

                        self.key_three_data[unit_key] = result['key_three_data']

                        # Collect Key Three-only units (missing from BeAScout)
                        if result['status'] == 'key_three_only':
                            missing_units.append(self._create_missing_unit_record(result['key_three_data']))
                
                # Add missing units to quality data with score=0, grade="N/A"
                self.quality_data['units_with_scores'].extend(missing_units)
                self.quality_data['total_units'] = len(self.quality_data['units_with_scores'])
                
                print(f"📊 Loaded Key Three data for {len(self.key_three_data)} units")
                print(f"📊 Added {len(missing_units)} units missing from BeAScout")
                
                # Store validation summary for executive summary
                self.validation_summary = validation_data['validation_summary']
            
            return True
        except Exception as e:
            print(f"❌ Failed to load data: {e}")
            return False
    
    def _create_missing_unit_record(self, key_three_data: Dict) -> Dict:
        """Create a unit record for Key Three-only units with score=0, grade=N/A"""
        return {
            'unit_key': key_three_data.get('unit_key', ''),
            'unit_type': key_three_data.get('unit_type', ''),
            'unit_number': key_three_data.get('unit_number', ''),
            'unit_town': key_three_data.get('unit_town', ''),
            'chartered_organization': key_three_data.get('chartered_organization', ''),
            'district': key_three_data.get('district', ''),
            'data_source': 'key_three_only',
            # Empty contact/meeting information since not present in BeAScout
            'meeting_location': '',
            'meeting_day': '',
            'meeting_time': '',
            'contact_email': '',
            'contact_person': '',
            'phone_number': '',
            'website': '',
            'description': '',
            'specialty': '',
            # Quality scoring for missing units
            'completeness_score': 0,
            'completeness_grade': 'N/A',
            'recommendations': [],
            # Mark as missing from BeAScout
            'missing_from_beascout': True
        }
    
    def create_quality_report(self, output_path: str = None, session_id: str = None) -> str:
        """
        Create comprehensive BeAScout Quality Report organized by district

        Args:
            output_path: Explicit output path (overrides session_id logic)
            session_id: Session ID for pipeline mode (generates weekly report path)

        Returns:
            Path to generated Excel file
        """
        if not self.quality_data:
            raise ValueError("No quality data loaded")

        # Generate timestamped filename if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            session_suffix = session_id if session_id else timestamp

            # Use weekly flag to determine report type and location
            if hasattr(self, '_weekly_mode') and self._weekly_mode:
                # Weekly pipeline mode: weekly reports directory
                output_path = f"data/output/reports/weekly/BeAScout_Weekly_Quality_Report_{session_suffix}.xlsx"
            else:
                # Standard commissioner report mode: main reports directory
                output_path = f"data/output/reports/BeAScout_Quality_Report_{session_suffix}.xlsx"
        
        # Ensure output directory exists
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        
        # Create Executive Summary first
        self._create_executive_summary_sheet(workbook)
        
        # Create district-specific sheets
        self._create_district_sheets(workbook)
        
        # Save workbook
        workbook.save(output_path)

        return output_path
    
    def _create_executive_summary_sheet(self, workbook):
        """Create executive summary sheet with quality metrics and legend"""
        ws = workbook.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "BeAScout Quality Report"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Council name and details
        ws['A2'] = "Heart of New England Council"
        ws['A2'].font = Font(size=14, bold=True)
        
        # Report generation info
        ws['A3'] = f"Generation Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(size=12)
        
        # Data sources - include Key Three spreadsheet name
        ws['A4'] = f"Data Sources: BeAScout.org (10-mile search radius per zip code) + JoinExploring.org (20-mile search radius per zip code) + {self.key_three_filename}"
        ws['A4'].font = Font(size=12)
        
        # Extract and format original scraping timestamp with date and time
        scraping_timestamp = self.quality_data.get('scraping_timestamp', '')
        if scraping_timestamp and scraping_timestamp != 'Unknown':
            # Parse ISO timestamp and format as readable date and time
            try:
                scraping_date = datetime.fromisoformat(scraping_timestamp.replace('Z', '+00:00'))
                formatted_date = scraping_date.strftime('%Y-%m-%d %H:%M:%S')
            except (ValueError, AttributeError):
                formatted_date = scraping_timestamp[:10] if len(scraping_timestamp) >= 10 else 'Date Unknown'
        else:
            formatted_date = 'Date Unknown'
        
        ws['A5'] = f"Last Complete BeAScout Data Retrieval: {formatted_date}"
        ws['A5'].font = Font(size=12)
        
        ws['A7'] = "This information within this document is to be used only for authorized purposes on behalf of the Scouting America."
        ws['A7'].font = Font(size=12, bold=True)
        ws['A8'] = "Disclosing, copying, or making any inappropriate use of this information is strictly prohibited."
        ws['A8'].font = Font(size=12, bold=True)

        # Quality metrics - remove extra spacing
        row = 10
        ws[f'A{row}'] = "QUALITY OVERVIEW"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        # Calculate metrics from quality data including missing units
        total_units = self.quality_data['total_units']
        
        # Calculate average score only for units with BeAScout data
        beascout_units = [u for u in self.quality_data['units_with_scores'] if not u.get('missing_from_beascout', False)]
        if beascout_units:
            avg_score = sum(u.get('completeness_score', 0) for u in beascout_units) / len(beascout_units)
        else:
            avg_score = 0
        
        # Count missing units
        missing_units_count = len([u for u in self.quality_data['units_with_scores'] if u.get('missing_from_beascout', False)])
        
        # Count units by district
        district_counts = {}
        town_counts = set()
        grade_counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0, 'N/A': 0}
        
        for unit in self.quality_data['units_with_scores']:
            district = unit.get('district', 'Unknown')
            district_counts[district] = district_counts.get(district, 0) + 1
            town_counts.add(unit.get('unit_town', 'Unknown'))
            grade = unit.get('completeness_grade', 'F')
            if grade in grade_counts:
                grade_counts[grade] += 1
        
        row += 1  # Reduced spacing
        metrics = [
            ("Total Units Analyzed", total_units),
            ("Units missing from BeAScout", missing_units_count),
            ("Units across Towns", f"{total_units} units across {len(town_counts)} towns"),
            ("Average Quality Score", f"{avg_score:.1f}%"),
            ("Quality Grade Distribution:", ""),
            ("  • Grade A (90%+)", f"{grade_counts['A']} units ({grade_counts['A']/total_units*100:.1f}%)"),
            ("  • Grade B (80-89%)", f"{grade_counts['B']} units ({grade_counts['B']/total_units*100:.1f}%)"),
            ("  • Grade C (70-79%)", f"{grade_counts['C']} units ({grade_counts['C']/total_units*100:.1f}%)"),
            ("  • Grade D (60-69%)", f"{grade_counts['D']} units ({grade_counts['D']/total_units*100:.1f}%)"),
            ("  • Grade F (<60%)", f"{grade_counts['F']} units ({grade_counts['F']/total_units*100:.1f}%)"),
            ("  • Grade N/A (Missing)", f"{grade_counts['N/A']} units ({grade_counts['N/A']/total_units*100:.1f}%)"),
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = str(value)
            if not metric.startswith('  •'):
                ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # District breakdown - reduced spacing
        row += 1  # Reduced spacing
        ws[f'A{row}'] = "UNITS BY DISTRICT"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 1  # Reduced spacing
        for district, count in sorted(district_counts.items()):
            ws[f'A{row}'] = district
            ws[f'B{row}'] = f"{count} units"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Quality Issue Legend with restructured format - reduced spacing
        row += 1  # Reduced spacing
        ws[f'A{row}'] = "QUALITY ISSUE LEGEND"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 1  # Reduced spacing
        # Required Information subsection
        ws[f'A{row}'] = "Required Information:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        required_items = [
            ("Meeting location", "Unit needs a physical meeting location with street address for parents and youth to find meetings"),
            ("Meeting day", "Unit needs to specify which day(s) of the week meetings are held"),
            ("Meeting time", "Unit needs to specify what time meetings start and end"),
            ("Contact email", "Unit needs a contact email address for inquiries and communication"),
        ]
        for issue, explanation in required_items:
            ws[f'A{row}'] = f"  • {issue}"
            ws[f'B{row}'] = explanation
            row += 1
        
        # Needs Improvement subsection  
        ws[f'A{row}'] = "Needs Improvement:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        quality_items = [
            ("Meeting location", "Provide physical meeting location in Unit Meeting Address field instead of Description field"),
            ("PO Box location", "Complement PO Box with physical meeting location so parents and youth can find meetings"),
            ("Personal email", "Use unit-specific email monitored by multiple leaders instead of personal email for continuity"),
        ]
        for issue, explanation in quality_items:
            ws[f'A{row}'] = f"  • {issue}"
            ws[f'B{row}'] = explanation
            row += 1
            
        # Recommended Information subsection
        ws[f'A{row}'] = "Recommended Information:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        recommended_items = [
            ("Contact person", "Adding a contact person name helps parents and youth know who to reach out to"),
            ("Phone number", "Phone contact provides immediate communication option for urgent questions"),
            ("Website", "Unit-specific website increases visibility and provides additional information for families"),
            ("Description", "Informative description helps attract new members by explaining unit activities and culture"),
        ]
        for issue, explanation in recommended_items:
            ws[f'A{row}'] = f"  • {issue}"
            ws[f'B{row}'] = explanation
            row += 1
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 60
    
    def _create_district_sheets(self, workbook):
        """Create separate sheets for each district with detailed unit information"""
        # Group units by district
        units_by_district = {}
        for unit in self.quality_data['units_with_scores']:
            district = unit.get('district', 'Unknown')
            if district not in units_by_district:
                units_by_district[district] = []
            units_by_district[district].append(unit)
        
        # Create sheet for each district
        for district_name in sorted(units_by_district.keys()):
            if district_name != 'Unknown':  # Skip unknown districts
                self._create_district_sheet(workbook, district_name, units_by_district[district_name])
    
    def _create_district_sheet(self, workbook, district_name: str, units: List[Dict]):
        """Create individual district sheet with detailed unit information"""
        ws = workbook.create_sheet(district_name)
        
        # Header section
        ws['A1'] = "BeAScout Quality Report"
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A2'] = "Heart of New England Council"
        ws['A2'].font = Font(size=14, bold=True)
        
        ws['A3'] = f"{district_name} District"
        ws['A3'].font = Font(size=14, bold=True)
        
        ws['A4'] = f"Generation Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'].font = Font(size=10)
        
        ws['A5'] = f"Data Sources: BeAScout.org (10-mile search radius per zip code) + JoinExploring.org (20-mile search radius per zip code) + {self.key_three_filename}"
        ws['A5'].font = Font(size=10)
        
        # Extract and format original scraping timestamp with date and time (same logic as Executive Summary)
        scraping_timestamp = self.quality_data.get('scraping_timestamp', '')
        if scraping_timestamp and scraping_timestamp != 'Unknown':
            try:
                scraping_date = datetime.fromisoformat(scraping_timestamp.replace('Z', '+00:00'))
                formatted_date = scraping_date.strftime('%Y-%m-%d %H:%M:%S')
            except (ValueError, AttributeError):
                formatted_date = scraping_timestamp[:10] if len(scraping_timestamp) >= 10 else 'Date Unknown'
        else:
            formatted_date = 'Date Unknown'
        
        ws['A6'] = f"Last Complete BeAScout Data Retrieval: {formatted_date}"
        ws['A6'].font = Font(size=10)
        
        # Count units and towns for this district
        towns = set(unit.get('unit_town', 'Unknown') for unit in units)
        ws['A7'] = f"{len(units)} Units across {len(towns)} towns"
        ws['A7'].font = Font(size=12, bold=True)
        
        # Create column headers (row 9) using ReportColumns definition
        for col_num, header in enumerate(ReportColumns.HEADERS, 1):
            cell = ws.cell(row=9, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Sort units by quality grade (worst first), then score, then unit identifier
        # Grade hierarchy: N/A (worst) < F < D < C < B < A (best)
        def sort_key(unit):
            grade = unit.get('completeness_grade', 'F')
            score = unit.get('completeness_score', 0)
            unit_key = unit.get('unit_key', '')
            
            # Map grades to sort order (lower numbers = worse grades = higher priority)
            grade_order = {'N/A': 0, 'F': 1, 'D': 2, 'C': 3, 'B': 4, 'A': 5}
            grade_priority = grade_order.get(grade, 1)  # Default to F priority
            
            return (grade_priority, score, unit_key)
        
        sorted_units = sorted(units, key=sort_key)
        
        # Add unit data starting from row 10
        for row_num, unit in enumerate(sorted_units, 10):
            self._populate_unit_row(ws, row_num, unit)
        
        # Auto-adjust column widths using ReportColumns definition
        for col_num, width in enumerate(ReportColumns.WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Freeze columns A through D (Unit Identifier through Zip Code) and header rows
        ws.freeze_panes = 'E10'
        
        # Apply professional formatting to all data cells
        self._apply_district_sheet_formatting(ws, len(sorted_units), ReportColumns.TOTAL_COLUMNS)
    
    def _populate_unit_row(self, ws, row_num: int, unit: Dict):
        """Populate a single unit row with all required information"""
        # Get Key Three information if available
        unit_key = unit.get('unit_key', '')
        key_three_info = self.key_three_data.get(unit_key, {})
        
        # Convert unit key to display format for report (remove leading zeros)
        display_unit_key = unit_key
        if unit_key:
            parts = unit_key.split()
            if len(parts) >= 3:
                unit_type = parts[0]
                unit_number_4digit = parts[1]
                town = ' '.join(parts[2:])
                display_number = UnitIdentifierNormalizer.get_display_unit_number(unit_number_4digit)
                display_unit_key = f"{unit_type} {display_number} {town}"
        key_three_members = key_three_info.get('key_three_members', [])
        
        # Categorize recommendations from quality tags stored in unit data during parsing
        # Use integrated quality_tags or fall back to recommendations field
        recommendations = unit.get('quality_tags', unit.get('recommendations', []))
        missing_info = [rec for rec in recommendations if rec.startswith('REQUIRED_')]
        quality_issues = [rec for rec in recommendations if rec.startswith('QUALITY_')]
        recommended_improvements = [rec for rec in recommendations if rec.startswith('RECOMMENDED_') or rec.startswith('CONTENT_')]
        
        # Get readable descriptions
        missing_descriptions = self.scorer.get_recommendation_descriptions(missing_info)
        quality_descriptions = self.scorer.get_recommendation_descriptions(quality_issues)
        recommended_descriptions = self.scorer.get_recommendation_descriptions(recommended_improvements)
        
        # Format Key Three member information (remove labels, multi-line, no unhelpful N/A)
        def format_key_three_member(member_info):
            if not member_info or member_info.get('fullname') == 'None':
                return "None"
            
            # Only include actual data, skip N/A values
            lines = []
            position = member_info.get('position', '')
            name = member_info.get('fullname', '')
            email = member_info.get('email', '')
            phone = member_info.get('phone', '')
            
            # Only add lines that have meaningful content
            if position and position != 'N/A':
                lines.append(position)
            if name and name != 'N/A':
                lines.append(name)
            # Skip address since not available in Key Three data
            if email and email != 'N/A':
                lines.append(email)
            if phone and phone != 'N/A':
                lines.append(phone)
            
            return '\n'.join(lines) if lines else 'None'
        
        key_three_1 = format_key_three_member(key_three_members[0] if len(key_three_members) > 0 else {})
        key_three_2 = format_key_three_member(key_three_members[1] if len(key_three_members) > 1 else {})
        key_three_3 = format_key_three_member(key_three_members[2] if len(key_three_members) > 2 else {})
        
        # Format meeting location with multi-line (facility\nstreet address\ncity/state/zip)
        meeting_location = unit.get('meeting_location', '')
        if meeting_location:
            # Smart parsing to avoid breaking multi-word town names
            # Only break on commas that aren't part of town names
            parts = []
            current_part = ''
            
            # Split on commas but be careful with multi-word towns
            for part in meeting_location.split(', '):
                part = part.strip()
                if part:
                    # Check if this looks like a state/zip (e.g., "MA 01234")
                    if re.match(r'^[A-Z]{2}\s+\d{5}', part):
                        # This is likely state/zip, add to parts
                        parts.append(part)
                    elif len(parts) == 0:
                        # First part is likely facility/building
                        parts.append(part)
                    elif re.match(r'^\d+\s+', part):
                        # Starts with number, likely street address
                        parts.append(part)
                    else:
                        # Could be continuation of address or town name
                        # If previous part looks incomplete, combine
                        if parts and not re.match(r'^\d+\s+.*\s+(Street|St|Road|Rd|Avenue|Ave|Lane|Ln|Drive|Dr|Way|Circle|Cir)', parts[-1], re.IGNORECASE):
                            parts[-1] = parts[-1] + ', ' + part
                        else:
                            parts.append(part)
            
            formatted_location = '\n'.join(parts) if parts else meeting_location
        else:
            formatted_location = ''
        
        # Get zip code for this unit's town
        unit_town = unit.get('unit_town', '')
        zip_code = get_zip_code_for_town(unit_town, self.town_zip_mapping)
        
        # Populate row data with Zip Code and reordered contact columns
        row_data = [
            display_unit_key,
            unit.get('chartered_organization', ''),
            unit_town,
            zip_code,                        # Populated Zip Code column
            unit.get('completeness_score', 0),
            unit.get('completeness_grade', 'F'),
            '\n'.join(missing_descriptions),  # Each issue on separate line
            '\n'.join(quality_descriptions),  # Each issue on separate line
            '\n'.join(recommended_descriptions),  # Each issue on separate line
            formatted_location,  # Multi-line meeting location
            unit.get('meeting_day', ''),
            unit.get('meeting_time', ''),
            unit.get('specialty', ''),       # Specialty field for Crews
            unit.get('contact_email', ''),   # Reordered: Contact Email first  
            unit.get('contact_person', ''),  # Then Contact Person
            unit.get('phone_number', ''),    # Then Contact Phone
            unit.get('website', ''),
            key_three_1,
            key_three_2,
            key_three_3
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            # Handle 0 values explicitly (don't convert to empty string)
            if value is None or value == '':
                cell.value = ""
            elif col_num == ReportColumns.QUALITY_SCORE:  # Quality Score column - use numeric format, left-aligned
                cell.value = float(value) if isinstance(value, (int, float)) else value
            elif col_num == ReportColumns.CONTACT_EMAIL and value:  # Contact Email column - make clickable
                email = str(value).strip()
                if email and '@' in email:
                    cell.hyperlink = f"mailto:{email}"
                    cell.value = email
                    cell.font = Font(color="0000FF", underline="single")  # Blue underlined text
                else:
                    cell.value = email
            elif col_num == ReportColumns.ZIP_CODE and value:  # Zip Code column - make clickable link to BeAScout/JoinExploring for this unit type
                zip_code = str(value).strip()
                if zip_code and len(zip_code) == 5:  # Valid 5-digit zip code
                    # Get unit type from row data for targeted search
                    unit_identifier = row_data[0] if row_data else ""  # First column is unit identifier
                    unit_type = ""
                    if unit_identifier:
                        if unit_identifier.startswith("Pack "):
                            unit_type = "pack"
                        elif unit_identifier.startswith("Troop "):
                            unit_type = "scoutsBSA" 
                        elif unit_identifier.startswith("Crew "):
                            unit_type = "crew"
                        elif unit_identifier.startswith("Ship "):
                            unit_type = "ship"
                        elif unit_identifier.startswith("Post "):
                            unit_type = "post"
                        elif unit_identifier.startswith("Club "):
                            unit_type = "club"
                    
                    if unit_type in ["post", "club"]:
                        # Create JoinExploring search URL for specific unit type
                        exploring_url = f"https://joinexploring.org/list/?zip={zip_code}&program[0]={unit_type}&miles=20"
                        cell.hyperlink = exploring_url
                    elif unit_type:
                        # Create BeAScout search URL for specific unit type
                        beascout_url = f"https://beascout.scouting.org/list/?zip={zip_code}&program[0]={unit_type}&miles=10"
                        cell.hyperlink = beascout_url
                    
                    cell.value = zip_code
                    cell.font = Font(color="0000FF", underline="single")  # Blue underlined text
                else:
                    cell.value = zip_code
            elif col_num == ReportColumns.UNIT_WEBSITE and value:  # Unit Website column - make clickable
                website = str(value).strip()
                if website:
                    # Add https:// if no protocol specified
                    if not website.startswith(('http://', 'https://')):
                        website_url = f"https://{website}"
                    else:
                        website_url = website
                    cell.hyperlink = website_url
                    cell.value = website
                    cell.font = Font(color="0000FF", underline="single")  # Blue underlined text
                else:
                    cell.value = website
            elif col_num in ReportColumns.KEY_THREE_COLUMNS and value:  # Key Three columns - no hyperlink (Excel limitation)
                text = str(value).strip()
                cell.value = text if text != 'None' else text
            else:
                cell.value = str(value)
            
            # Set alignment - left-aligned for Quality Score, default for others
            if col_num == ReportColumns.QUALITY_SCORE:
                cell.alignment = Alignment(horizontal='left', wrap_text=True, vertical='top')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Color code quality grades including N/A for missing units
            if col_num == ReportColumns.QUALITY_GRADE:  # Quality Grade column
                grade = str(value)
                if grade == 'A':
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif grade == 'B':
                    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                elif grade == 'C':
                    cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                elif grade == 'D':
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                elif grade == 'F':
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                elif grade == 'N/A':
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for missing units
    
    def _apply_district_sheet_formatting(self, ws, num_data_rows: int, num_columns: int):
        """Apply professional formatting to district sheet with borders"""
        from openpyxl.styles import Border, Side
        
        # Define border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply formatting to header row (row 9)
        for col_num in range(1, num_columns + 1):
            cell = ws.cell(row=9, column=col_num)
            cell.border = thin_border
            
        # Define column category fill colors (very light colors)
        quality_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")      # Very light yellow (F-H)
        beascout_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")     # Very light blue (I-O)  
        key_three_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")    # Very light green (P-R)
        
        # Apply formatting to data rows (starting from row 10)
        for row_num in range(10, 10 + num_data_rows):
            for col_num in range(1, num_columns + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                
                # Apply category fill colors using ReportColumns categories
                if col_num in ReportColumns.QUALITY_COLUMNS[2:]:  # Quality Issues, Missing Info, Recommended Improvements
                    cell.fill = quality_fill
                elif col_num in ReportColumns.MEETING_COLUMNS:   # Meeting and Contact Information
                    cell.fill = beascout_fill
                elif col_num in ReportColumns.KEY_THREE_COLUMNS:  # Key Three Contacts
                    cell.fill = key_three_fill

def main():
    """Generate BeAScout Quality Report organized by districts with CLI support"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Generate comprehensive BeAScout Quality Report with Key Three integration',
        epilog='''
Examples:
  # Independent mode - general reports directory
  python generate_commissioner_report.py
  → data/output/reports/BeAScout_Quality_Report_20250920_143025.xlsx

  # Pipeline mode - weekly reports directory
  python generate_commissioner_report.py --session-id 20250920_143025
  → data/output/reports/weekly/BeAScout_Weekly_Quality_Report_20250920_143025.xlsx

  # Use anonymized test data
  python generate_commissioner_report.py --key-three tests/reference/key_three/anonymized_key_three.json

  # Use custom data sources with explicit output
  python generate_commissioner_report.py --quality-data data/raw/custom_units.json --output-dir data/reports/

Output Path Logic:
  • Independent mode (no --session-id): data/output/reports/BeAScout_Quality_Report_[TIMESTAMP].xlsx
  • Pipeline mode (with --session-id): data/output/reports/weekly/BeAScout_Weekly_Quality_Report_[SESSION_ID].xlsx
  • Explicit --output-dir: Uses specified directory with appropriate filename format

Pipeline Dependencies:
  1. process_full_dataset.py → --quality-data
  2. three_way_validator.py  → --validation-file
  3. anonymize_key_three.py  → --key-three (test mode)
        ''',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument('--quality-data',
                       default='data/raw/all_units_comprehensive_scored.json',
                       help='Path to quality-scored units JSON [default: data/raw/all_units_comprehensive_scored.json] [generated by: process_full_dataset.py]')
    parser.add_argument('--validation-file',
                       default='data/output/enhanced_three_way_validation_results.json', 
                       help='Path to validation results JSON [default: data/output/enhanced_three_way_validation_results.json] [generated by: three_way_validator.py]')
    parser.add_argument('--key-three',
                       help='Path to Key Three data (JSON/Excel) [default: uses Key Three data from validation file] [anonymized version: anonymize_key_three.py]')
    parser.add_argument('--output-dir',
                       help='Output directory for reports [default: auto-determined by mode]')
    parser.add_argument('--session-id',
                       help='Session ID for pipeline mode')
    parser.add_argument('--weekly', action='store_true',
                       help='Generate weekly report format (used by weekly pipeline)')
    parser.add_argument('--scraped-session',
                       help='Scraped session ID for accurate data timestamp display')

    # Add additional session management arguments (note: --session-id already exists above)
    parser.add_argument('--log', action='store_true',
                       help='Direct stdout/stderr to log file')

    args = parser.parse_args()

    # Create session manager from arguments (using existing session-id)
    session_manager = SessionManager(session_id=args.session_id, session_type='pipeline')

    # Use session logging context manager with terminal_terse mode
    with session_logging(session_manager, "generate_commissioner_report",
                        log_enabled=args.log, verbose=args.log, terminal_terse=True):

        # All detailed processing goes to log, show terse summary on terminal
        session_manager.terse_print("📋 Generating BeAScout Quality Report")

        # Extract just the filename (not full path) for display in reports
        import os
        key_three_display_name = os.path.basename(args.key_three) if args.key_three else None

        generator = BeAScoutQualityReportGenerator(key_three_display_name, args.scraped_session)

        # Set weekly mode flag based on command line argument
        generator._weekly_mode = args.weekly

        # Load quality and Key Three data with custom paths (detailed output goes to log)
        if not generator.load_quality_data(args.quality_data, args.validation_file):
            session_manager.terse_print("❌ Failed to load quality data")
            return

        # Generate report with conditional output path logic (detailed output goes to log)
        if args.output_dir:
            # Explicit output directory specified
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            session_suffix = args.session_id if args.session_id else timestamp

            if args.weekly:
                output_path = f"{args.output_dir.rstrip('/')}/BeAScout_Weekly_Quality_Report_{session_suffix}.xlsx"
            else:
                output_path = f"{args.output_dir.rstrip('/')}/BeAScout_Quality_Report_{session_suffix}.xlsx"
        else:
            output_path = None  # Use method's conditional logic

        report_path = generator.create_quality_report(output_path, args.session_id)

        # Get statistics for terse terminal display
        total_units = generator.quality_data['total_units']
        avg_score = generator.quality_data.get('average_score', 0)

        # Count districts
        districts = set()
        for unit in generator.quality_data['units_with_scores']:
            district = unit.get('district', 'Unknown')
            if district != 'Unknown':
                districts.add(district)

        # Show terse summary on terminal
        session_manager.terse_print(f"✅ BeAScout Quality Report Generated Successfully!")
        session_manager.terse_print(f"📁 Report saved to: {report_path}")
        session_manager.terse_print(f"📊 {total_units} total units, {len(districts)} districts, avg score: {avg_score}%")

        return report_path

if __name__ == "__main__":
    main()