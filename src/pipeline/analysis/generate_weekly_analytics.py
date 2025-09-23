#!/usr/bin/env python3
"""
Weekly Analytics Generator for BeAScout Quality Reports
Extracts statistics from Excel reports and compares with previous week
Generates simple JSON metadata for week-over-week analysis
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

# Add project root to path
project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root))

class WeeklyAnalyticsGenerator:
    """Generates weekly analytics from BeAScout Quality Reports"""

    def __init__(self):
        self.current_report_path = None
        self.current_analytics = None
        self.previous_analytics = None

    def find_latest_report(self, reports_dir: Path = None) -> Optional[Path]:
        """Find the most recent weekly quality report"""
        if reports_dir is None:
            reports_dir = project_root / "data" / "output" / "reports" / "weekly"

        if not reports_dir.exists():
            print(f"❌ Reports directory not found: {reports_dir}")
            return None

        # Find Excel files matching weekly report pattern
        report_files = list(reports_dir.glob("BeAScout_Weekly_Quality_Report_*.xlsx"))

        if not report_files:
            print(f"❌ No weekly quality reports found in: {reports_dir}")
            return None

        # Return most recent by modification time
        latest_report = max(report_files, key=lambda p: p.stat().st_mtime)
        print(f"📊 Found latest report: {latest_report.name}")

        return latest_report

    def find_previous_analytics(self, current_timestamp: str, reports_dir: Path, baseline_file: str = None) -> Optional[Dict]:
        """Find and load previous week's analytics for comparison"""
        if baseline_file:
            # Use explicit baseline file
            baseline_path = Path(baseline_file)

            # If baseline_file is just a filename, look in reports_dir
            if not baseline_path.is_absolute():
                baseline_path = reports_dir / baseline_file

            if not baseline_path.exists():
                print(f"❌ Baseline file not found: {baseline_path}")
                return None

            try:
                with open(baseline_path, 'r', encoding='utf-8') as f:
                    baseline_data = json.load(f)

                print(f"📈 Using explicit baseline: {baseline_path.name}")
                # Add baseline metadata for transparency
                baseline_data['_baseline_metadata'] = {
                    'baseline_filename': baseline_path.name,
                    'baseline_timestamp': baseline_data.get('report_metadata', {}).get('timestamp', 'unknown'),
                    'baseline_date': baseline_data.get('report_metadata', {}).get('report_date', 'unknown')
                }
                return baseline_data

            except Exception as e:
                print(f"❌ Could not load baseline file {baseline_path.name}: {e}")
                return None

        # Auto-detect previous analytics (original behavior)
        analytics_files = list(reports_dir.glob("BeAScout_Weekly_Quality_Report_*.json"))

        if not analytics_files:
            print("ℹ️  No previous analytics found - this may be the first weekly report")
            return None

        # Filter out current timestamp and find most recent
        previous_files = [f for f in analytics_files if current_timestamp not in f.name]

        if not previous_files:
            print("ℹ️  No previous analytics found - this may be the first weekly report")
            return None

        previous_file = max(previous_files, key=lambda p: p.stat().st_mtime)

        try:
            with open(previous_file, 'r', encoding='utf-8') as f:
                previous_data = json.load(f)

            print(f"📈 Auto-detected previous analytics: {previous_file.name}")
            # Add baseline metadata for transparency
            previous_data['_baseline_metadata'] = {
                'baseline_filename': previous_file.name,
                'baseline_timestamp': previous_data.get('report_metadata', {}).get('timestamp', 'unknown'),
                'baseline_date': previous_data.get('report_metadata', {}).get('report_date', 'unknown')
            }
            return previous_data

        except Exception as e:
            print(f"⚠️  Could not load previous analytics from {previous_file.name}: {e}")
            return None

    def extract_executive_summary_stats(self, excel_path: Path) -> Dict:
        """Extract key statistics from Executive Summary sheet"""
        try:
            # Load workbook and get Executive Summary sheet
            workbook = load_workbook(excel_path, data_only=True)

            if "Executive Summary" not in workbook.sheetnames:
                raise ValueError("Executive Summary sheet not found in workbook")

            ws = workbook["Executive Summary"]

            # Initialize stats
            stats = {
                "total_units": None,
                "average_quality_score": None,
                "grade_distribution": {},
                "units_missing_from_beascout": None
            }

            # Scan through rows looking for key metrics
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue

                    cell_text = str(cell.value).strip()

                    # Extract total units
                    if "Total Units Analyzed" in cell_text:
                        # Look for number in adjacent cell
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value is not None:
                            try:
                                stats["total_units"] = int(next_cell.value)
                            except (ValueError, TypeError):
                                pass

                    # Extract average quality score
                    elif "Average Quality Score" in cell_text:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value is not None:
                            value_str = str(next_cell.value)
                            # Extract number before % sign
                            if "%" in value_str:
                                try:
                                    stats["average_quality_score"] = float(value_str.replace("%", ""))
                                except (ValueError, TypeError):
                                    pass

                    # Extract units missing from BeAScout
                    elif "Units missing from BeAScout" in cell_text:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value is not None:
                            try:
                                stats["units_missing_from_beascout"] = int(next_cell.value)
                            except (ValueError, TypeError):
                                pass

                    # Extract grade distribution
                    elif "Grade A (90%+)" in cell_text:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value is not None:
                            stats["grade_distribution"]["A"] = self._extract_count_and_percentage(next_cell.value)

                    elif "Grade B (80-89%)" in cell_text:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value is not None:
                            stats["grade_distribution"]["B"] = self._extract_count_and_percentage(next_cell.value)

                    elif "Grade C (70-79%)" in cell_text:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value is not None:
                            stats["grade_distribution"]["C"] = self._extract_count_and_percentage(next_cell.value)

                    elif "Grade D (60-69%)" in cell_text:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value is not None:
                            stats["grade_distribution"]["D"] = self._extract_count_and_percentage(next_cell.value)

                    elif "Grade F (<60%)" in cell_text:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value is not None:
                            stats["grade_distribution"]["F"] = self._extract_count_and_percentage(next_cell.value)

                    elif "Grade N/A (Missing)" in cell_text:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value is not None:
                            stats["grade_distribution"]["N/A"] = self._extract_count_and_percentage(next_cell.value)

            print(f"✅ Extracted executive summary statistics")
            return stats

        except Exception as e:
            print(f"❌ Error extracting executive summary: {e}")
            return {}

    def _extract_count_and_percentage(self, cell_value) -> Dict:
        """Extract count and percentage from cell like '25 units (15.2%)'"""
        try:
            text = str(cell_value).strip()

            # Extract count (number before 'units')
            count = None
            if "units" in text:
                count_part = text.split("units")[0].strip()
                try:
                    count = int(count_part)
                except ValueError:
                    pass

            # Extract percentage (number between parentheses)
            percentage = None
            if "(" in text and "%" in text:
                pct_part = text.split("(")[1].split("%")[0].strip()
                try:
                    percentage = float(pct_part)
                except ValueError:
                    pass

            return {"count": count, "percentage": percentage}

        except Exception:
            return {"count": None, "percentage": None}

    def extract_unit_scores(self, excel_path: Path) -> Dict[str, float]:
        """Extract unit scores from district sheets for comparison"""
        try:
            workbook = load_workbook(excel_path, data_only=True)
            unit_scores = {}

            # Process each district sheet (skip Executive Summary)
            for sheet_name in workbook.sheetnames:
                if sheet_name == "Executive Summary":
                    continue

                ws = workbook[sheet_name]

                # Find header row (should contain "Unit Identifier" and "Quality Score")
                header_row = None
                unit_col = None
                score_col = None

                for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=15), 1):
                    for col_num, cell in enumerate(row, 1):
                        if cell.value and "Unit Identifier" in str(cell.value):
                            header_row = row_num
                            unit_col = col_num
                        elif cell.value and "Quality Score" in str(cell.value):
                            score_col = col_num

                    if header_row and unit_col and score_col:
                        break

                if not (header_row and unit_col and score_col):
                    print(f"⚠️  Could not find headers in district sheet: {sheet_name}")
                    continue

                # Extract unit data starting from row after header
                for row in ws.iter_rows(min_row=header_row + 1):
                    unit_cell = row[unit_col - 1]
                    score_cell = row[score_col - 1]

                    if unit_cell.value and score_cell.value is not None:
                        unit_identifier = str(unit_cell.value).strip()
                        try:
                            score = float(score_cell.value)
                            unit_scores[unit_identifier] = score
                        except (ValueError, TypeError):
                            pass

            print(f"✅ Extracted {len(unit_scores)} unit scores")
            return unit_scores

        except Exception as e:
            print(f"❌ Error extracting unit scores: {e}")
            return {}

    def calculate_week_over_week_changes(self) -> Dict:
        """Calculate changes between current and previous week"""
        if not self.previous_analytics:
            return {"first_week": True, "changes": {}}

        current_exec = self.current_analytics.get("executive_summary", {})
        previous_exec = self.previous_analytics.get("executive_summary", {})

        changes = {}

        # Calculate total units change
        current_total = current_exec.get("total_units")
        previous_total = previous_exec.get("total_units")
        if current_total is not None and previous_total is not None:
            changes["total_units"] = {
                "current": current_total,
                "previous": previous_total,
                "change": current_total - previous_total
            }

        # Calculate average quality score change
        current_avg = current_exec.get("average_quality_score")
        previous_avg = previous_exec.get("average_quality_score")
        if current_avg is not None and previous_avg is not None:
            changes["average_quality_score"] = {
                "current": current_avg,
                "previous": previous_avg,
                "change": round(current_avg - previous_avg, 1)
            }

        # Calculate grade distribution changes
        current_grades = current_exec.get("grade_distribution", {})
        previous_grades = previous_exec.get("grade_distribution", {})

        grade_changes = {}
        for grade in ["A", "B", "C", "D", "F", "N/A"]:
            current_data = current_grades.get(grade, {})
            previous_data = previous_grades.get(grade, {})

            current_count = current_data.get("count")
            previous_count = previous_data.get("count")

            if current_count is not None and previous_count is not None:
                grade_changes[grade] = {
                    "current": current_count,
                    "previous": previous_count,
                    "change": current_count - previous_count
                }

        if grade_changes:
            changes["grade_distribution"] = grade_changes

        return {"first_week": False, "changes": changes}

    def calculate_unit_score_changes(self) -> Dict:
        """Calculate unit-level score improvements and declines"""
        if not self.previous_analytics:
            return {"first_week": True, "unit_changes": []}

        current_scores = self.current_analytics.get("unit_scores", {})
        previous_scores = self.previous_analytics.get("unit_scores", {})

        unit_changes = []

        # Find units present in both weeks
        common_units = set(current_scores.keys()) & set(previous_scores.keys())

        for unit_key in common_units:
            current_score = current_scores[unit_key]
            previous_score = previous_scores[unit_key]

            if current_score != previous_score:
                change = round(current_score - previous_score, 1)
                unit_changes.append({
                    "unit_identifier": unit_key,
                    "current_score": current_score,
                    "previous_score": previous_score,
                    "change": change,
                    "improvement": change > 0
                })

        # Sort by absolute change (largest changes first)
        unit_changes.sort(key=lambda x: abs(x["change"]), reverse=True)

        return {"first_week": False, "unit_changes": unit_changes}

    def generate_analytics(self, excel_path: Path, output_path: Path = None, baseline_file: str = None) -> Path:
        """Generate complete analytics JSON file"""
        self.current_report_path = excel_path

        # Extract timestamp from filename for output naming
        timestamp_match = excel_path.stem.split("_")[-2:]  # Get last two parts (date_time)
        if len(timestamp_match) == 2:
            timestamp = f"{timestamp_match[0]}_{timestamp_match[1]}"
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Determine output path
        if output_path is None:
            output_dir = excel_path.parent
            output_path = output_dir / f"BeAScout_Weekly_Quality_Report_{timestamp}.json"

        # Load previous analytics for comparison
        self.previous_analytics = self.find_previous_analytics(timestamp, excel_path.parent, baseline_file)

        # Extract current week's data
        print("📊 Extracting executive summary statistics...")
        executive_summary = self.extract_executive_summary_stats(excel_path)

        print("📊 Extracting unit scores...")
        unit_scores = self.extract_unit_scores(excel_path)

        # Build complete analytics structure
        self.current_analytics = {
            "report_metadata": {
                "report_date": timestamp[:8],  # YYYYMMDD
                "timestamp": timestamp,
                "excel_filename": excel_path.name,
                "generation_time": datetime.now().isoformat()
            },
            "executive_summary": executive_summary,
            "unit_scores": unit_scores
        }

        # Calculate week-over-week changes
        print("📈 Calculating week-over-week changes...")
        week_changes = self.calculate_week_over_week_changes()
        unit_changes = self.calculate_unit_score_changes()

        # Add comparison data
        comparison_data = {
            "executive_summary_changes": week_changes,
            "unit_score_changes": unit_changes
        }

        # Add baseline metadata if available
        if self.previous_analytics and '_baseline_metadata' in self.previous_analytics:
            comparison_data["baseline_metadata"] = self.previous_analytics['_baseline_metadata']

        self.current_analytics["weekly_comparison"] = comparison_data

        # Save analytics file
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(self.current_analytics, f, indent=2)

            print(f"✅ Analytics saved: {output_path}")
            return output_path

        except Exception as e:
            print(f"❌ Error saving analytics: {e}")
            raise

def main():
    """CLI interface for weekly analytics generation"""
    parser = argparse.ArgumentParser(
        description="Generate weekly analytics from BeAScout Quality Report",
        epilog="""
Examples:
  # Generate analytics from latest report
  python generate_weekly_analytics.py

  # Generate analytics from specific report
  python generate_weekly_analytics.py --excel-file path/to/report.xlsx

  # Use explicit baseline for comparison
  python generate_weekly_analytics.py --baseline BeAScout_Weekly_Quality_Report_20250904_154530.json

  # Specify output location
  python generate_weekly_analytics.py --output path/to/analytics.json

This script extracts key statistics from the Excel report and compares
with baseline data to generate week-over-week analytics.
        """,
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    parser.add_argument('--excel-file', help='Path to Excel report file [default: find latest in weekly reports directory]')
    parser.add_argument('--output', help='Output path for analytics JSON [default: same directory as Excel file]')
    parser.add_argument('--baseline', help='Baseline analytics file for comparison [default: auto-detect most recent]')

    args = parser.parse_args()

    generator = WeeklyAnalyticsGenerator()

    # Determine input Excel file
    if args.excel_file:
        excel_path = Path(args.excel_file)
        if not excel_path.exists():
            print(f"❌ Excel file not found: {excel_path}")
            sys.exit(1)
    else:
        excel_path = generator.find_latest_report()
        if not excel_path:
            print("❌ No weekly quality reports found")
            sys.exit(1)

    # Determine output path
    output_path = Path(args.output) if args.output else None

    # Generate analytics
    try:
        result_path = generator.generate_analytics(excel_path, output_path, args.baseline)
        print(f"\n✅ Weekly analytics generated successfully!")
        print(f"📁 Analytics file: {result_path}")

        # Display summary of changes
        if generator.current_analytics:
            comparison = generator.current_analytics.get("weekly_comparison", {})
            exec_changes = comparison.get("executive_summary_changes", {})

            if not exec_changes.get("first_week", True):
                print(f"\n📈 Week-over-Week Summary:")
                changes = exec_changes.get("changes", {})

                if "total_units" in changes:
                    total_change = changes["total_units"]["change"]
                    sign = "+" if total_change >= 0 else ""
                    print(f"   Total Units: {changes['total_units']['current']} ({sign}{total_change})")

                if "average_quality_score" in changes:
                    avg_change = changes["average_quality_score"]["change"]
                    sign = "+" if avg_change >= 0 else ""
                    print(f"   Average Score: {changes['average_quality_score']['current']}% ({sign}{avg_change}%)")

                # Show top unit changes
                unit_comparison = comparison.get("unit_score_changes", {})
                unit_changes = unit_comparison.get("unit_changes", [])

                if unit_changes:
                    improvements = [u for u in unit_changes if u["improvement"]][:3]
                    declines = [u for u in unit_changes if not u["improvement"]][:3]

                    if improvements:
                        print(f"\n🏆 Top Unit Improvements:")
                        for unit in improvements:
                            print(f"   {unit['unit_identifier']}: {unit['previous_score']}% → {unit['current_score']}% (+{unit['change']}%)")

                    if declines:
                        print(f"\n⚠️  Unit Score Declines:")
                        for unit in declines:
                            print(f"   {unit['unit_identifier']}: {unit['previous_score']}% → {unit['current_score']}% ({unit['change']}%)")
            else:
                print(f"\nℹ️  This appears to be the first weekly report - no comparison data available")

    except Exception as e:
        print(f"❌ Error generating analytics: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()